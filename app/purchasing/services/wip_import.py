import csv
import re
import uuid
from collections import Counter
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from audit.services import record_audit
from imports.models import RawFile
from imports.services.storage import DuplicateRawFile, _checksum
from master_data.models import SKU, Supplier

from ..models import (
    POWIPImportBatch,
    POWIPImportIssue,
    PurchaseOrder,
    PurchaseOrderLine,
    StagedPOWIPRow,
)


PARSER_VERSION = "po-wip-v1-jul31"
CUTOVER_DATE = date(2026, 7, 31)
MIGRATION_PO_DATE = date(2026, 7, 31)
SUPPLIER_CODE = "VOBIA-VENDOR"
SUPPLIER_NAME = "Vobia Vendor"
REQUIRED_HEADERS = ("NO PO", "SKU Induk", "SKU", "Nama Barang", "WIP")
PO_PATTERN = re.compile(r"^PO-VOB-(?P<month>\d{2})/(?P<year>\d{2})-(?P<sequence>\d{3})$")


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _quantity(value):
    if value in (None, "") or isinstance(value, bool):
        raise InvalidOperation
    number = Decimal(str(value).replace(",", "").strip())
    if not number.is_finite() or number <= 0 or number != number.to_integral_value():
        raise InvalidOperation
    return number


def _read(path, file_format):
    if file_format == "csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(8192)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
            values = list(csv.reader(handle, dialect=dialect))
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            values = [list(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
    if not values:
        raise ValidationError("File PO WIP kosong.")
    headers = [_text(value) for value in values[0]]
    rows = []
    for row_number, row in enumerate(values[1:], start=2):
        if not any(value not in (None, "") for value in row):
            continue
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        rows.append((row_number, dict(zip(headers, padded, strict=False))))
    return headers, rows


def _need_month(po_number):
    match = PO_PATTERN.fullmatch(po_number)
    if not match:
        raise ValueError("Format No. PO wajib PO-VOB-MM/YY-NNN.")
    month = int(match.group("month"))
    year = 2000 + int(match.group("year"))
    try:
        result = date(year, month, 1)
    except ValueError as exc:
        raise ValueError("Month pada No. PO tidak valid.") from exc
    if result > date(CUTOVER_DATE.year, CUTOVER_DATE.month, 1):
        raise ValueError("PO WIP tidak boleh memiliki Need Month setelah July 2026.")
    return result


def _issue(batch, row, severity, code, field_name, message, blocking):
    return POWIPImportIssue(
        batch=batch,
        staged_row=row,
        severity=severity,
        code=code,
        field_name=field_name,
        message=message,
        is_blocking=blocking,
    )


@transaction.atomic
def parse_po_wip_batch(batch):
    batch = POWIPImportBatch.objects.select_for_update().select_related("raw_file").get(pk=batch.pk)
    path = Path(settings.PRIVATE_UPLOAD_ROOT) / batch.raw_file.storage_path
    try:
        headers, source_rows = _read(path, batch.raw_file.detected_format)
    except Exception as exc:
        POWIPImportIssue.objects.create(
            batch=batch,
            severity=POWIPImportIssue.Severity.ERROR,
            code="FILE_PARSE_ERROR",
            message=str(exc),
            is_blocking=True,
        )
        batch.status = POWIPImportBatch.Status.BLOCKED
        batch.blocking_issue_count = 1
        batch.previewed_at = timezone.now()
        batch.save()
        return batch

    missing_headers = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing_headers:
        POWIPImportIssue.objects.create(
            batch=batch,
            severity=POWIPImportIssue.Severity.ERROR,
            code="MISSING_HEADERS",
            field_name="header",
            message="Header wajib tidak ditemukan: " + ", ".join(missing_headers),
            is_blocking=True,
        )
        batch.status = POWIPImportBatch.Status.BLOCKED
        batch.blocking_issue_count = 1
        batch.previewed_at = timezone.now()
        batch.quality_summary = {"headers": headers}
        batch.save()
        return batch

    pending = []
    staged_rows = []
    master = {
        item.sku: item
        for item in SKU.objects.filter(is_active=True).select_related("product_variant__product")
    }
    existing_po_numbers = set(PurchaseOrder.objects.values_list("po_number", flat=True))
    total_qty = Decimal("0")
    total_value = Decimal("0")
    name_mismatch_count = 0

    if POWIPImportBatch.objects.filter(status=POWIPImportBatch.Status.COMMITTED).exclude(pk=batch.pk).exists():
        pending.append(
            (
                None,
                "ERROR",
                "PO_WIP_ALREADY_COMMITTED",
                "batch",
                "Migrasi PO WIP immutable sudah pernah di-commit.",
                True,
            )
        )

    for row_number, raw in source_rows:
        po_number = _text(raw.get("NO PO"))
        parent_sku = _text(raw.get("SKU Induk"))
        sku_text = _text(raw.get("SKU"))
        product_name = _text(raw.get("Nama Barang"))
        sku = master.get(sku_text)
        quantity = None
        cogs = sku.current_master_cogs if sku else None
        row_issues = []

        try:
            _need_month(po_number)
        except ValueError as exc:
            row_issues.append(("ERROR", "INVALID_PO_NUMBER", "NO PO", str(exc), True))
        if po_number in existing_po_numbers:
            row_issues.append(("ERROR", "PO_NUMBER_ALREADY_EXISTS", "NO PO", f"{po_number} sudah ada di ERP.", True))
        try:
            quantity = _quantity(raw.get("WIP"))
        except (InvalidOperation, ValueError):
            row_issues.append(("ERROR", "INVALID_WIP_QTY", "WIP", "WIP wajib bilangan bulat positif.", True))
        if not sku_text:
            row_issues.append(("ERROR", "MISSING_SKU", "SKU", "SKU wajib diisi.", True))
        elif sku is None:
            row_issues.append(("ERROR", "UNKNOWN_SKU", "SKU", f"SKU {sku_text} tidak ada di Master Data aktif.", True))
        else:
            master_parent = (sku.product_variant.product.parent_sku or "").strip()
            if parent_sku.upper() != master_parent.upper():
                row_issues.append(
                    (
                        "ERROR",
                        "PARENT_SKU_MISMATCH",
                        "SKU Induk",
                        f"Parent SKU file {parent_sku or '—'} tidak cocok dengan master {master_parent or '—'}.",
                        True,
                    )
                )
            if cogs is None:
                row_issues.append(("ERROR", "MISSING_COGS", "SKU", "Current master COGS kosong; FIFO inbound tidak dapat dinilai.", True))
            if product_name.lower() != sku.product_variant.product.name.strip().lower():
                name_mismatch_count += 1

        if quantity is not None:
            total_qty += quantity
            if cogs is not None:
                total_value += quantity * cogs

        staged = StagedPOWIPRow.objects.create(
            batch=batch,
            row_number=row_number,
            po_number=po_number,
            parent_sku=parent_sku,
            sku_text=sku_text,
            product_name_source=product_name,
            outstanding_qty=quantity,
            sku=sku,
            proposed_cogs_snapshot=cogs,
            original_data={header: _text(raw.get(header)) for header in headers},
        )
        staged_rows.append(staged)
        pending.extend((staged, *issue) for issue in row_issues)

    counts = Counter((row.po_number, row.sku_text) for row in staged_rows if row.po_number and row.sku_text)
    for row in staged_rows:
        if counts[(row.po_number, row.sku_text)] > 1:
            pending.append(
                (
                    row,
                    "ERROR",
                    "DUPLICATE_PO_SKU",
                    "SKU",
                    f"{row.po_number} + {row.sku_text} muncul lebih dari sekali.",
                    True,
                )
            )

    pending.extend(
        [
            (
                None,
                "WARNING",
                "MIGRATION_DATE_AND_QTY_ASSUMPTION",
                "batch",
                "Atas keputusan Adit, seluruh PO memakai Open Date 31 July 2026; PO Qty = WIP; Received sebelum cutoff = 0.",
                False,
            ),
            (
                None,
                "WARNING",
                "MIGRATION_QC_ASSUMPTION",
                "batch",
                "Seluruh WIP diperlakukan sebagai QC Passed sebelum cutover agar eligible untuk physical Inbound ERP.",
                False,
            ),
            (
                None,
                "WARNING",
                "MASTER_COGS_PROPOSED_SNAPSHOT",
                "batch",
                "Current master COGS digunakan sebagai usulan snapshot FIFO PO WIP dan baru dibekukan saat approval.",
                False,
            ),
        ]
    )

    issue_models = [_issue(batch, *issue) for issue in pending]
    POWIPImportIssue.objects.bulk_create(issue_models)
    blocked_ids = {issue[0].id for issue in pending if issue[0] is not None and issue[-1]}
    for row in staged_rows:
        row.proposed_action = (
            StagedPOWIPRow.ProposedAction.BLOCKED
            if row.id in blocked_ids
            else StagedPOWIPRow.ProposedAction.NEW
        )
    StagedPOWIPRow.objects.bulk_update(staged_rows, ["proposed_action"])

    blocking = sum(1 for issue in issue_models if issue.is_blocking)
    warnings = sum(1 for issue in issue_models if issue.severity == POWIPImportIssue.Severity.WARNING)
    po_numbers = sorted({row.po_number for row in staged_rows if row.po_number})
    batch.status = POWIPImportBatch.Status.BLOCKED if blocking else POWIPImportBatch.Status.READY
    batch.parser_version = PARSER_VERSION
    batch.total_rows = len(staged_rows)
    batch.ready_rows = sum(1 for row in staged_rows if row.proposed_action == StagedPOWIPRow.ProposedAction.NEW)
    batch.po_count = len(po_numbers)
    batch.total_outstanding_qty = total_qty
    batch.blocking_issue_count = blocking
    batch.warning_count = warnings
    batch.previewed_at = timezone.now()
    batch.quality_summary = {
        "headers": headers,
        "supplier": SUPPLIER_NAME,
        "migration_po_date": str(MIGRATION_PO_DATE),
        "migration_cutoff_date": str(CUTOVER_DATE),
        "po_numbers": po_numbers,
        "total_proposed_cogs_value": str(total_value),
        "name_mismatch_count": name_mismatch_count,
        "duplicate_po_sku_count": sum(1 for count in counts.values() if count > 1),
        "received_before_cutover_qty": "0",
        "qc_passed_basis": "WIP qty",
        "required_arrival": None,
    }
    batch.save()
    return batch


def create_po_wip_import(uploaded_file, actor):
    extension = Path(uploaded_file.name).suffix.lower().lstrip(".")
    checksum = _checksum(uploaded_file)
    duplicate = RawFile.objects.filter(
        dataset_type=RawFile.DatasetType.PO_WIP,
        checksum_sha256=checksum,
    ).first()
    if duplicate:
        raise DuplicateRawFile(duplicate)

    current = timezone.localdate()
    relative = (
        Path("purchasing")
        / "po_wip"
        / str(current.year)
        / f"{current.month:02d}"
        / f"{uuid.uuid4()}.{extension}"
    )
    absolute = Path(settings.PRIVATE_UPLOAD_ROOT) / relative
    absolute.parent.mkdir(parents=True, exist_ok=True)
    with absolute.open("wb") as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)
    uploaded_file.seek(0)

    try:
        with transaction.atomic():
            raw = RawFile.objects.create(
                dataset_type=RawFile.DatasetType.PO_WIP,
                original_filename=Path(uploaded_file.name).name,
                storage_path=str(relative),
                checksum_sha256=checksum,
                byte_size=uploaded_file.size,
                detected_format=extension,
                uploaded_by=actor,
                source_metadata={
                    "dataset": "po_wip_outstanding",
                    "migration_cutoff_date": str(CUTOVER_DATE),
                    "migration_po_date": str(MIGRATION_PO_DATE),
                    "supplier_assumption": SUPPLIER_NAME,
                },
            )
            batch = POWIPImportBatch.objects.create(raw_file=raw)
            record_audit(
                actor=actor,
                action="po_wip_import_uploaded",
                entity_type="purchasing.powipimportbatch",
                entity_id=batch.id,
                metadata={
                    "filename": raw.original_filename,
                    "checksum_sha256": checksum,
                    "byte_size": raw.byte_size,
                },
            )
    except Exception:
        absolute.unlink(missing_ok=True)
        raise
    return parse_po_wip_batch(batch)


@transaction.atomic
def approve_po_wip_import(batch_id, actor):
    batch = (
        POWIPImportBatch.objects.select_for_update()
        .select_related("raw_file")
        .get(pk=batch_id)
    )
    if not batch.can_approve or batch.issues.filter(is_blocking=True).exists():
        raise ValidationError("Batch PO WIP belum siap atau masih memiliki blocking issue.")
    if POWIPImportBatch.objects.filter(status=POWIPImportBatch.Status.COMMITTED).exclude(pk=batch.pk).exists():
        raise ValidationError("Migrasi PO WIP immutable sudah pernah di-commit.")

    rows = list(batch.staged_rows.select_related("sku__product_variant__product").order_by("row_number"))
    if len(rows) != batch.ready_rows or any(row.proposed_action != StagedPOWIPRow.ProposedAction.NEW for row in rows):
        raise ValidationError("Cakupan staging berubah sejak preview; unggah ulang file sumber.")

    po_numbers = {row.po_number for row in rows}
    if PurchaseOrder.objects.filter(po_number__in=po_numbers).exists():
        raise ValidationError("Satu atau lebih No. PO sudah ada sejak preview; commit dibatalkan.")
    for row in rows:
        row.sku.refresh_from_db()
        if not row.sku.is_active:
            raise ValidationError(f"SKU {row.sku_text} sudah tidak aktif sejak preview.")
        if row.sku.current_master_cogs != row.proposed_cogs_snapshot:
            raise ValidationError(f"COGS {row.sku_text} berubah sejak preview; unggah ulang untuk review snapshot terbaru.")

    supplier, _ = Supplier.objects.get_or_create(
        code=SUPPLIER_CODE,
        defaults={"name": SUPPLIER_NAME, "is_active": True},
    )
    evidence = f"{batch.raw_file.original_filename} · SHA256 {batch.raw_file.checksum_sha256}"
    # Midday Jakarta preserves the intended 31 July calendar date even when
    # ORM values are inspected in UTC during tests/audit exports.
    migration_datetime = timezone.make_aware(datetime.combine(MIGRATION_PO_DATE, time(hour=12)))
    created_pos = []
    line_count = 0

    for po_number in sorted(po_numbers):
        po_rows = [row for row in rows if row.po_number == po_number]
        po = PurchaseOrder(
            po_number=po_number,
            supplier=supplier,
            need_month=_need_month(po_number),
            required_arrival=None,
            source=PurchaseOrder.Source.LEGACY_WIP,
            status=PurchaseOrder.Status.RELEASED,
            notes=(
                "Migrasi outstanding PO WIP per 31 July 2026. PO Qty = WIP; Received before cutover = 0; "
                "QC Passed before cutover = WIP; Required Arrival kosong; COGS memakai approved master snapshot."
            ),
            created_by=actor,
            released_by=actor,
            released_at=migration_datetime,
            migration_cutoff_date=CUTOVER_DATE,
            migration_evidence_reference=evidence,
        )
        po.full_clean()
        po.save()
        PurchaseOrder.objects.filter(pk=po.pk).update(created_at=migration_datetime)
        po.refresh_from_db()
        created_pos.append(po)
        for row in po_rows:
            line = PurchaseOrderLine(
                po=po,
                sku=row.sku,
                ordered_qty=row.outstanding_qty,
                cogs_snapshot=row.proposed_cogs_snapshot,
                received_before_cutover_qty=Decimal("0"),
                qc_passed_before_cutover_qty=row.outstanding_qty,
            )
            line.full_clean()
            line.save()
            line_count += 1

    now = timezone.now()
    batch.status = POWIPImportBatch.Status.COMMITTED
    batch.approved_by = actor
    batch.approved_at = now
    batch.committed_at = now
    batch.save(update_fields=["status", "approved_by", "approved_at", "committed_at"])
    counts = {
        "purchase_orders": len(created_pos),
        "lines": line_count,
        "outstanding_qty": str(sum((row.outstanding_qty for row in rows), Decimal("0"))),
        "inventory_movements": 0,
        "fifo_layers": 0,
    }
    record_audit(
        actor=actor,
        action="po_wip_import_committed",
        entity_type="purchasing.powipimportbatch",
        entity_id=batch.id,
        after_values=counts,
        metadata={
            "checksum_sha256": batch.raw_file.checksum_sha256,
            "supplier": supplier.name,
            "migration_po_date": str(MIGRATION_PO_DATE),
            "migration_cutoff_date": str(CUTOVER_DATE),
        },
    )
    return batch, counts
