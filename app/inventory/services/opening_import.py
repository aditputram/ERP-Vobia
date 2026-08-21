import csv
import math
import uuid
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from audit.services import record_audit
from imports.models import RawFile
from imports.services.storage import DuplicateRawFile, _checksum
from master_data.models import SKU

from ..models import FIFOOpeningImportBatch, FIFOOpeningImportIssue, FIFOOpeningSnapshot, StagedFIFOOpeningRow
from .fifo import CUTOVER_DATE, post_opening


PARSER_VERSION = "fifo-opening-v2-jul31"
REQUIRED_HEADERS = ("SKU", "Opening Qty", "Frozen Unit COGS", "Cutover Date")


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal(value):
    if value in (None, "") or isinstance(value, bool):
        raise InvalidOperation
    number = Decimal(str(value).replace(",", "").strip())
    if not number.is_finite():
        raise InvalidOperation
    return number


def _date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value)
        return converted.date() if isinstance(converted, datetime) else converted
    text = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError("Cutover Date tidak dikenali.")


def _read(path, file_format):
    if file_format == "csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(8192)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
            values = list(csv.reader(handle, dialect=dialect))
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = next((sheet for sheet in workbook.worksheets if sheet.title.strip() == "FIFO Opening"), workbook.active)
            values = [list(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
    if not values:
        raise ValidationError("File FIFO Opening kosong.")
    headers = [_text(value) for value in values[0]]
    rows = []
    for row_number, row in enumerate(values[1:], start=2):
        if not any(value not in (None, "") for value in row):
            continue
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        rows.append((row_number, dict(zip(headers, padded, strict=False))))
    return headers, rows


def _issue(batch, row, severity, code, field_name, message, blocking):
    return FIFOOpeningImportIssue(
        batch=batch, staged_row=row, severity=severity, code=code,
        field_name=field_name, message=message, is_blocking=blocking,
    )


@transaction.atomic
def parse_opening_batch(batch):
    batch = FIFOOpeningImportBatch.objects.select_for_update().select_related("raw_file").get(pk=batch.pk)
    path = Path(settings.PRIVATE_UPLOAD_ROOT) / batch.raw_file.storage_path
    try:
        headers, source_rows = _read(path, batch.raw_file.detected_format)
    except Exception as exc:
        FIFOOpeningImportIssue.objects.create(batch=batch, severity="ERROR", code="FILE_PARSE_ERROR", message=str(exc), is_blocking=True)
        batch.status = FIFOOpeningImportBatch.Status.BLOCKED
        batch.blocking_issue_count = 1
        batch.previewed_at = timezone.now()
        batch.save()
        return batch

    missing_headers = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing_headers:
        FIFOOpeningImportIssue.objects.create(
            batch=batch, severity="ERROR", code="MISSING_HEADERS", field_name="header",
            message="Header wajib tidak ditemukan: " + ", ".join(missing_headers), is_blocking=True,
        )
        batch.status = FIFOOpeningImportBatch.Status.BLOCKED
        batch.blocking_issue_count = 1
        batch.previewed_at = timezone.now()
        batch.quality_summary = {"headers": headers}
        batch.save()
        return batch

    master = {item.sku: item for item in SKU.objects.filter(is_active=True)}
    pending = []
    staged_rows = []
    total_qty = Decimal("0")
    total_value = Decimal("0")
    negative_count = 0
    zero_count = 0

    if FIFOOpeningSnapshot.objects.exists():
        pending.append((None, "ERROR", "OPENING_ALREADY_POSTED", "batch", "FIFO Opening immutable sudah pernah diposting.", True))

    for row_number, raw in source_rows:
        sku_text = _text(raw.get("SKU"))
        sku = master.get(sku_text)
        quantity = None
        cost = None
        cutover = None
        row_issues = []
        try:
            quantity = _decimal(raw.get("Opening Qty"))
            if quantity != quantity.to_integral_value():
                raise InvalidOperation
        except (InvalidOperation, ValueError):
            row_issues.append(("ERROR", "INVALID_OPENING_QTY", "Opening Qty", "Opening Qty wajib bilangan bulat, termasuk nol/negatif.", True))
        try:
            cost = _decimal(raw.get("Frozen Unit COGS"))
            if cost < 0:
                raise InvalidOperation
        except (InvalidOperation, ValueError):
            row_issues.append(("ERROR", "INVALID_FROZEN_COGS", "Frozen Unit COGS", "Frozen Unit COGS wajib angka nonnegative.", True))
        try:
            cutover = _date(raw.get("Cutover Date"))
            if cutover != CUTOVER_DATE:
                row_issues.append(("ERROR", "INVALID_CUTOVER_DATE", "Cutover Date", "Cutover Date wajib 31 July 2026 (saldo akhir hari).", True))
        except ValueError as exc:
            row_issues.append(("ERROR", "INVALID_CUTOVER_DATE", "Cutover Date", str(exc), True))
        if not sku_text:
            row_issues.append(("ERROR", "MISSING_SKU", "SKU", "SKU wajib diisi.", True))
        elif sku is None:
            row_issues.append(("ERROR", "UNKNOWN_SKU", "SKU", f"SKU {sku_text} tidak ada di Master Data aktif.", True))
        if quantity is not None and quantity < 0:
            negative_count += 1
            row_issues.append(("WARNING", "NEGATIVE_OPENING", "Opening Qty", "Negative opening dipertahankan dan akan menjadi inventory exception.", False))
        if quantity == 0:
            zero_count += 1
        if quantity is not None and cost is not None:
            total_qty += quantity
            total_value += quantity * cost
        staged = StagedFIFOOpeningRow.objects.create(
            batch=batch, row_number=row_number, sku_text=sku_text, sku=sku,
            opening_qty=quantity, frozen_unit_cogs=cost, cutover_date=cutover,
            source_layer_key=_text(raw.get("Layer Key")),
            original_data={header: _text(raw.get(header)) for header in headers},
        )
        staged_rows.append(staged)
        pending.extend((staged, *issue) for issue in row_issues)

    counts = Counter(row.sku_text for row in staged_rows if row.sku_text)
    for row in staged_rows:
        if counts[row.sku_text] > 1:
            pending.append((row, "ERROR", "DUPLICATE_SKU", "SKU", f"SKU {row.sku_text} muncul lebih dari sekali.", True))
    source_skus = set(counts)
    missing_master = sorted(set(master) - source_skus)
    if missing_master:
        pending.append((None, "ERROR", "MASTER_SKU_MISSING_FROM_OPENING", "SKU", f"{len(missing_master)} SKU Master tidak ada di file Opening. Contoh: {', '.join(missing_master[:10])}", True))

    issue_models = [_issue(batch, *issue) for issue in pending]
    FIFOOpeningImportIssue.objects.bulk_create(issue_models)
    blocked_ids = {issue[0].id for issue in pending if issue[0] is not None and issue[-1]}
    for row in staged_rows:
        row.proposed_action = StagedFIFOOpeningRow.ProposedAction.BLOCKED if row.id in blocked_ids else StagedFIFOOpeningRow.ProposedAction.NEW
    StagedFIFOOpeningRow.objects.bulk_update(staged_rows, ["proposed_action"])
    blocking = sum(1 for issue in issue_models if issue.is_blocking)
    warnings = sum(1 for issue in issue_models if issue.severity == "WARNING")
    batch.status = FIFOOpeningImportBatch.Status.BLOCKED if blocking else FIFOOpeningImportBatch.Status.READY
    batch.parser_version = PARSER_VERSION
    batch.total_rows = len(staged_rows)
    batch.ready_rows = sum(1 for row in staged_rows if row.proposed_action == StagedFIFOOpeningRow.ProposedAction.NEW)
    batch.blocking_issue_count = blocking
    batch.warning_count = warnings
    batch.previewed_at = timezone.now()
    batch.quality_summary = {
        "headers": headers, "master_sku_count": len(master), "negative_opening_count": negative_count,
        "zero_opening_count": zero_count, "total_opening_qty": str(total_qty),
        "total_opening_value": str(total_value), "duplicate_sku_count": sum(1 for count in counts.values() if count > 1),
    }
    batch.save()
    return batch


def create_opening_import(uploaded_file, actor):
    extension = Path(uploaded_file.name).suffix.lower().lstrip(".")
    checksum = _checksum(uploaded_file)
    duplicate = RawFile.objects.filter(dataset_type=RawFile.DatasetType.FIFO_OPENING, checksum_sha256=checksum).first()
    if duplicate:
        raise DuplicateRawFile(duplicate)
    current = timezone.localdate()
    relative = Path("inventory") / "fifo_opening" / str(current.year) / f"{current.month:02d}" / f"{uuid.uuid4()}.{extension}"
    absolute = Path(settings.PRIVATE_UPLOAD_ROOT) / relative
    absolute.parent.mkdir(parents=True, exist_ok=True)
    with absolute.open("wb") as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)
    uploaded_file.seek(0)
    try:
        with transaction.atomic():
            raw = RawFile.objects.create(
                dataset_type=RawFile.DatasetType.FIFO_OPENING, original_filename=Path(uploaded_file.name).name,
                storage_path=str(relative), checksum_sha256=checksum, byte_size=uploaded_file.size,
                detected_format=extension, uploaded_by=actor,
                source_metadata={"workbook": "Vobia MD 2026", "sheet": "FIFO Opening", "cutover_date": str(CUTOVER_DATE)},
            )
            batch = FIFOOpeningImportBatch.objects.create(raw_file=raw)
            record_audit(actor=actor, action="fifo_opening_import_uploaded", entity_type="inventory.fifoopeningimportbatch", entity_id=batch.id, metadata={"checksum_sha256": checksum})
    except Exception:
        absolute.unlink(missing_ok=True)
        raise
    return parse_opening_batch(batch)


@transaction.atomic
def approve_opening_import(batch_id, actor):
    batch = FIFOOpeningImportBatch.objects.select_for_update().get(pk=batch_id)
    if not batch.can_approve or batch.issues.filter(is_blocking=True).exists():
        raise ValidationError("Batch FIFO Opening belum siap atau masih memiliki blocking issue.")
    if FIFOOpeningSnapshot.objects.exists():
        raise ValidationError("FIFO Opening immutable sudah pernah diposting.")
    rows = list(batch.staged_rows.select_related("sku").order_by("row_number"))
    expected = set(SKU.objects.filter(is_active=True).values_list("id", flat=True))
    actual = {row.sku_id for row in rows}
    if expected != actual:
        raise ValidationError("Cakupan SKU berubah sejak preview; parse ulang dari export terbaru.")
    for row in rows:
        post_opening(sku=row.sku, quantity=row.opening_qty, unit_cost=row.frozen_unit_cogs, actor=actor, reason="Bulk FIFO cutover opening · approved import")
    now = timezone.now()
    batch.status = FIFOOpeningImportBatch.Status.COMMITTED
    batch.approved_by = actor
    batch.approved_at = now
    batch.committed_at = now
    batch.save(update_fields=["status", "approved_by", "approved_at", "committed_at"])
    counts = {
        "snapshots": len(rows),
        "positive_layers": sum(1 for row in rows if row.opening_qty > 0),
        "zero_openings": sum(1 for row in rows if row.opening_qty == 0),
        "negative_exceptions": sum(1 for row in rows if row.opening_qty < 0),
    }
    record_audit(actor=actor, action="fifo_opening_import_committed", entity_type="inventory.fifoopeningimportbatch", entity_id=batch.id, after_values=counts, metadata={"checksum_sha256": batch.raw_file.checksum_sha256})
    return batch, counts
