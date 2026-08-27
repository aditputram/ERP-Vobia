from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import uuid

from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.test import RequestFactory, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import User
from inventory.models import InventoryException, InventoryMovement
from imports.models import RawFile
from master_data.models import Category, MarketplaceProductMapping, Product, ProductStatus, ProductVariant, SKU, Subcategory
from merchandising.models import MerchandisingMonthlySnapshot, MerchandisingSnapshotBatch
from traffic.models import TrafficImportBatch, TrafficProductMetric

from .models import SalesOrder, SalesOrderLine, SalesPlan, SalesPlanSKU, SalesPlanningScenario
from .views import _sales_planning_totals, _save_sales_projection_preview


class SalesReportRouteTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(username="vobiasuperadmin", password="strong-test-password")
        self.client.force_login(self.user)

    def _planning_product(self, code="PLAN-PRODUCT"):
        status, _ = ProductStatus.objects.get_or_create(code="PLAN-ACTIVE", defaults={"name": "Active"})
        category, _ = Category.objects.get_or_create(code="PLAN-CATEGORY", defaults={"name": "Planning Category"})
        product = Product.objects.create(code=code, name=f"Product {code}", status=status, category=category)
        variant = ProductVariant.objects.create(product=product, name="Default")
        sku = SKU.objects.create(sku=f"SKU-{code}", product_variant=variant, current_retail_price=Decimal("100000"))
        return product, sku

    def test_sales_report_routes_render(self):
        for name in ("sales:planning_builder", "sales:product_performance", "sales:pareto", "sales:transactions", "sales:input_transaction"):
            with self.subTest(name=name):
                response = self.client.get(reverse(name))
                self.assertEqual(response.status_code, 200)

    def test_sales_plan_summary_aggregates_saved_targets_with_cascading_filters(self):
        product, sku = self._planning_product("SUMMARY-PANTS")
        second_sku = SKU.objects.create(sku="SUMMARY-PANTS-L", product_variant=sku.product_variant, size="L")
        other, other_sku = self._planning_product("SUMMARY-SHIRT")
        other.status = ProductStatus.objects.create(code="SUMMARY-SEASONAL", name="Seasonal")
        other.category = Category.objects.create(code="SUMMARY-SHIRT", name="Shirts")
        other.subcategory = Subcategory.objects.create(category=other.category, code="LINEN", name="Linen")
        other.save()
        product.subcategory = Subcategory.objects.create(category=product.category, code="DENIM", name="Denim")
        product.save()
        scenario = SalesPlanningScenario.objects.create(name="Summary Draft", start_month=date(2026, 9, 1), end_month=date(2026, 10, 1), created_by=self.user)
        approved = SalesPlanningScenario.objects.create(name="Summary Approved", start_month=date(2026, 10, 1), end_month=date(2026, 10, 1), created_by=self.user, status=SalesPlanningScenario.Status.APPROVED)
        september = SalesPlan.objects.create(scenario=scenario, product=product, month=date(2026, 9, 1), quantity_target=999)
        october = SalesPlan.objects.create(scenario=scenario, product=product, month=date(2026, 10, 1))
        other_plan = SalesPlan.objects.create(scenario=approved, product=other, month=date(2026, 10, 1))
        for plan, item, qty, gross in [(september, sku, 10, 1000), (september, second_sku, 2, 200), (october, sku, 3, 300), (other_plan, other_sku, 5, 1000)]:
            SalesPlanSKU.objects.create(plan=plan, sku=item, quantity_target=qty, gross_sales_target=gross)
        context = {"scenario": scenario.id, "month": "2026-09", "draft_month": ["2026-09", "2026-10"], "draft_metric": "qty", "draft_grain": "parent_sku"}
        page = self.client.get(reverse("sales:planning_builder"), context)
        summary = page.context["plan_summary"]
        self.assertEqual(summary["rows"], [
            {"month": date(2026, 9, 1), "qty": 12, "gross": Decimal("1200")},
            {"month": date(2026, 10, 1), "qty": 8, "gross": Decimal("1300")},
        ])
        self.assertEqual((summary["qty"], summary["gross"]), (20, Decimal("2500")))
        self.assertLess(page.content.decode().index('id="sales-plan-summary"'), page.content.decode().index('01 · SCENARIO'))
        filtered = self.client.get(reverse("sales:planning_builder"), {
            **context, "summary_start": "2026-10", "summary_end": "2026-10", "summary_product": [str(other.id)],
        })
        self.assertEqual(filtered.context["plan_summary"]["qty"], 5)
        self.assertEqual(filtered.context["target_totals"]["qty"], 15)  # Summary does not filter the draft.
        self.assertEqual(filtered.context["selected_draft_months"], [date(2026, 9, 1), date(2026, 10, 1)])
        cascade = self.client.get(reverse("sales:planning_builder"), {
            **context, "summary_status": str(product.status_id), "summary_category": str(other.category_id),
            "summary_subcategory": str(other.subcategory_id), "summary_product": str(other.id),
        }).context["plan_summary"]
        self.assertEqual(cascade["qty"], 15)
        self.assertEqual([option["value"] for option in cascade["filters"][1]["options"]], [str(product.category_id)])
        self.assertEqual([option["value"] for option in cascade["filters"][2]["options"]], [str(product.subcategory_id)])
        self.assertEqual([option["value"] for option in cascade["filters"][3]["options"]], [str(product.id)])
        self.assertTrue(all(not row["selected"] for row in cascade["filters"][1:]))
        multi = self.client.get(reverse("sales:planning_builder"), {**context, "summary_product": [str(product.id), str(other.id), str(product.id)]}).context["plan_summary"]
        self.assertEqual(multi["qty"], 20)
        self.assertEqual(SalesPlanSKU.objects.count(), 4)

    def test_sales_plan_summary_validates_month_range_and_empty_data(self):
        for filters in ({"summary_start": "2026-11", "summary_end": "2026-09"}, {"summary_start": "bad-month"}):
            page = self.client.get(reverse("sales:planning_builder"), filters)
            self.assertEqual(page.status_code, 200)
            self.assertTrue(page.context["plan_summary"]["error"])
            self.assertEqual(page.context["plan_summary"]["rows"], [])
        empty = self.client.get(reverse("sales:planning_builder"), {"summary_start": "2027-01", "summary_end": "2027-12"})
        self.assertContains(empty, "Belum ada target tersimpan untuk filter ini")
        self.assertEqual(empty.context["plan_summary"]["qty"], 0)
        self.assertFalse(empty.context["plan_summary"]["error"])

    def test_sales_draft_matrix_months_metrics_and_scoped_save(self):
        first, sku = self._planning_product("MATRIX-A")
        second, second_sku = self._planning_product("MATRIX-B")
        Product.objects.filter(pk__in=[first.pk, second.pk]).update(parent_sku="MATRIX-PARENT")
        scenario = SalesPlanningScenario.objects.create(
            name="Matrix months", start_month=date(2026, 9, 1), end_month=date(2026, 12, 1), created_by=self.user,
        )
        targets = []
        for product, item, month, qty in [(first, sku, 9, 10), (first, sku, 10, 20), (second, second_sku, 10, 5), (first, sku, 11, 30)]:
            plan = SalesPlan.objects.create(scenario=scenario, product=product, month=date(2026, month, 1), quantity_target=qty, gross_sales_target=qty * 100000)
            targets.append(SalesPlanSKU.objects.create(plan=plan, sku=item, quantity_target=qty, gross_sales_target=qty * 100000))
        filters = {"scenario": scenario.id, "month": "2026-12", "draft_month": ["2026-10", "2026-09", "2026-10"], "draft_metric": ["qty"]}
        page = self.client.get(reverse("sales:planning_builder"), filters)
        self.assertEqual(page.context["selected_draft_months"], [date(2026, 9, 1), date(2026, 10, 1)])
        self.assertEqual(len(page.context["rows"]), 2)
        self.assertEqual(len(page.context["draft_parent_rows"]), 1)
        self.assertEqual(page.context["target_totals"]["qty"], 35)
        self.assertEqual(page.context["target_totals"]["gross"], 3500000)
        self.assertEqual(page.context["target_totals"]["sku_count"], 2)
        self.assertEqual([cell["qty"] for cell in page.context["target_totals"]["targets"]], [10, 25])
        self.assertEqual(page.context["draft_history_headers"][0]["month"], date(2026, 6, 1))
        self.assertIsNone(page.context["rows"][1]["targets"][0]["target"])
        self.assertContains(page, "Target September 2026")
        self.assertContains(page, "Target Oktober 2026")
        self.assertNotIn('<th>Gross Sales</th>', page.content.decode().split('id="draft-projection"', 1)[1])
        self.assertContains(page, '<tfoot>', count=3)  # Summary plus both draft grains.
        gross_page = self.client.get(reverse("sales:planning_builder"), {**filters, "draft_metric": "gross", "draft_grain": "parent_sku"})
        self.assertNotContains(gross_page, '<th>Sales Qty</th>')
        self.assertContains(gross_page, f'type="hidden" name="qty_{targets[0].id}" value="10"')
        self.assertEqual(gross_page.context["selected_draft_grain"], "parent_sku")
        both = self.client.get(reverse("sales:planning_builder"), {**filters, "draft_metric": ["qty", "gross"]})
        self.assertContains(both, '<th>Sales Qty</th>')
        self.assertContains(both, '<th>Gross Sales</th>')
        payload = {**filters, "form_name": "projection", "draft_grain": "parent_sku",
                   **{f"qty_{target.id}": target.quantity_target + 1 for target in targets[:3]}}
        saved = self.client.post(reverse("sales:planning_builder"), payload, follow=True)
        self.assertEqual(saved.context["selected_draft_months"], page.context["selected_draft_months"])
        self.assertEqual(saved.context["target_totals"]["qty"], 38)
        self.assertEqual(saved.context["selected_draft_metrics"], ["qty"])
        self.assertEqual(saved.context["selected_draft_grain"], "parent_sku")
        for target, expected in zip(targets, [11, 21, 6, 30]):
            target.refresh_from_db()
            self.assertEqual(target.quantity_target, expected)
            target.plan.refresh_from_db()
            self.assertEqual(target.plan.quantity_target, expected)
        # No field may silently zero an omitted target or edit an unselected month.
        stale = {**payload}
        del stale[f"qty_{targets[0].id}"]
        self.assertContains(self.client.post(reverse("sales:planning_builder"), stale, follow=True), "Isi Draft telah berubah")
        invalid = {**payload, f"qty_{targets[0].id}": "88", f"qty_{targets[1].id}": "-1"}
        self.client.post(reverse("sales:planning_builder"), invalid)
        targets[0].refresh_from_db()
        self.assertEqual(targets[0].quantity_target, 11)
        outside = {**payload, f"qty_{targets[3].id}": "99"}
        self.assertContains(self.client.post(reverse("sales:planning_builder"), outside, follow=True), "Isi Draft telah berubah")
        targets[3].refresh_from_db()
        self.assertEqual(targets[3].quantity_target, 30)

    def test_sales_planning_totals_sum_all_rows_and_weight_growth(self):
        months = [date(2026, 6, 1), date(2026, 7, 1), date(2026, 8, 1)]
        groups = [
            {"sku_count": 6, "target_qty": 20, "target_gross": Decimal("2000"),
             "history": [{"qty": 10, "gross": Decimal("1000")} for _ in months]},
            {"sku_count": 7, "target_qty": 30, "target_gross": Decimal("6000"),
             "history": [{"qty": 30, "gross": Decimal("6000")} for _ in months]},
        ]
        totals = _sales_planning_totals(groups, months)
        self.assertEqual(totals["sku_count"], 13)
        self.assertEqual(totals["qty"], 50)
        self.assertEqual(totals["gross"], Decimal("8000"))
        self.assertEqual(totals["growth_pct"], Decimal("25"))
        self.assertEqual(totals["history"], [
            {"month": month, "qty": 40, "gross": Decimal("7000")} for month in months
        ])
        empty = _sales_planning_totals([], months)
        self.assertEqual(empty["qty"], 0)
        self.assertEqual(empty["gross"], 0)
        self.assertIsNone(empty["growth_pct"])

    def test_sales_route_keeps_sales_navigation_when_other_tab_selected_operation(self):
        session = self.client.session
        session["active_module"] = "operation"
        session.save()

        response = self.client.get(reverse("sales:planning_builder"))

        self.assertContains(response, reverse("sales:planning_builder"))
        self.assertNotContains(response, reverse("merchandising:planning_builder"))
        self.assertContains(response, ">Sales<")

    def test_sales_planning_scenario_builds_monthly_projections_then_approves(self):
        product, sku = self._planning_product()
        order = SalesOrder.objects.create(
            source=SalesOrder.Source.SHOPEE,
            source_label="Shopee",
            order_number="PLAN-ACTUAL-1",
            order_datetime=timezone.make_aware(datetime(2026, 9, 5, 10, 0)),
            order_date=date(2026, 9, 5),
            current_status="Selesai",
            source_status="Selesai",
            is_final=True,
            first_seen_batch_id=uuid.uuid4(),
            latest_batch_id=uuid.uuid4(),
        )
        SalesOrderLine.objects.create(
            order=order,
            sku=sku,
            sku_code_snapshot="PLAN-SKU",
            product_name_snapshot="Planning Product",
            quantity=2,
            net_unit_price=Decimal("90000"),
            retail_price_snapshot=Decimal("100000"),
            total_gross_sales=Decimal("200000"),
            total_net_sales=Decimal("180000"),
        )
        created = self.client.post(reverse("sales:planning_builder"), {
            "form_name": "scenario",
            "name": "Sales Plan Sep-Oct 2026",
            "start_month": "2026-09",
            "end_month": "2026-10",
        }, follow=True)
        scenario = SalesPlanningScenario.objects.get(name="Sales Plan Sep-Oct 2026")
        self.assertEqual(scenario.status, SalesPlanningScenario.Status.DRAFT)
        self.assertContains(created, "SCENARIO LIBRARY")

        plan = SalesPlan.objects.create(
            scenario=scenario,
            month=date(2026, 9, 1),
            product=product,
            gross_sales_target=Decimal("250000"),
            quantity_target=2,
        )
        target = SalesPlanSKU.objects.create(
            plan=plan,
            sku=sku,
            gross_sales_target=Decimal("250000"),
            quantity_target=2,
        )
        payload = {
            "form_name": "projection",
            "scenario": str(scenario.id),
            "month": "2026-09",
            f"qty_{target.id}": "3",
        }

        saved = self.client.post(reverse("sales:planning_builder"), payload, follow=True)
        plan.refresh_from_db()
        self.assertEqual(plan.gross_sales_target, Decimal("300000"))
        product_row = saved.context["rows"][0]
        self.assertEqual(product_row["product"], product)
        self.assertEqual(product_row["actual"]["gross"], Decimal("200000"))
        self.assertEqual(product_row["gross_gap"], Decimal("-100000"))

        incomplete = self.client.post(reverse("sales:planning_builder"), {
            "form_name": "approval", "scenario": str(scenario.id), "month": "2026-09",
        }, follow=True)
        scenario.refresh_from_db()
        self.assertEqual(scenario.status, SalesPlanningScenario.Status.DRAFT)
        self.assertContains(incomplete, "Projection belum lengkap untuk: Oct 2026")

        october_plan = SalesPlan.objects.create(
            scenario=scenario,
            month=date(2026, 10, 1),
            product=product,
            gross_sales_target=Decimal("400000"),
            quantity_target=4,
        )
        SalesPlanSKU.objects.create(
            plan=october_plan,
            sku=sku,
            gross_sales_target=Decimal("400000"),
            quantity_target=4,
        )

        approved = self.client.post(reverse("sales:planning_builder"), {
            "form_name": "approval", "scenario": str(scenario.id), "month": "2026-10",
        }, follow=True)
        scenario.refresh_from_db()
        self.assertEqual(scenario.status, SalesPlanningScenario.Status.APPROVED)
        self.assertEqual(scenario.approved_by, self.user)
        self.assertContains(approved, "approved dan seluruh projection dikunci")

    def test_sales_projection_builder_previews_three_months_then_saves_selected_product(self):
        product, sku = self._planning_product("BUILDER-PRODUCT")
        product.parent_sku = "PARENT-BUILDER"
        product.save(update_fields=["parent_sku"])
        SKU.objects.create(
            sku="SKU-BUILDER-PRODUCT-L",
            product_variant=sku.product_variant,
            size="L",
        )
        order = SalesOrder.objects.create(
            source=SalesOrder.Source.SHOPEE,
            source_label="Shopee",
            order_number="BUILDER-AUGUST-1",
            order_datetime=timezone.make_aware(datetime(2026, 8, 5, 10, 0)),
            order_date=date(2026, 8, 5),
            current_status="Selesai",
            source_status="Selesai",
            is_final=True,
            first_seen_batch_id=uuid.uuid4(),
            latest_batch_id=uuid.uuid4(),
        )
        SalesOrderLine.objects.create(
            order=order,
            sku=sku,
            sku_code_snapshot="BUILDER-SKU",
            product_name_snapshot="Builder Product",
            quantity=10,
            net_unit_price=Decimal("90000"),
            retail_price_snapshot=Decimal("100000"),
            total_gross_sales=Decimal("1000000"),
            total_net_sales=Decimal("900000"),
        )
        batch = MerchandisingSnapshotBatch.objects.create(
            source_workbook_id="sales-builder-tiering",
            source_file_name="sales-builder-tiering.xlsx",
            source_sha256="b" * 64,
            imported_by=self.user,
            row_count=2,
            is_active=True,
        )
        for source_row, snapshot_month in enumerate((date(2026, 7, 1), date(2026, 8, 1)), start=1):
            MerchandisingMonthlySnapshot.objects.create(
                batch=batch,
                sku=sku,
                source_row=source_row,
                month=snapshot_month,
                status_snapshot="Active",
                product_snapshot=product.name,
                category_snapshot=product.category.name,
                retail_price_snapshot=Decimal("100000"),
                ending_qty=Decimal("100") if snapshot_month.month == 7 else Decimal("0"),
            )
        scenario = SalesPlanningScenario.objects.create(
            name="Sales Builder September 2026",
            start_month=date(2026, 9, 1),
            end_month=date(2026, 9, 1),
            created_by=self.user,
        )
        payload = {
            "form_name": "builder",
            "action": "preview",
            "scenario": str(scenario.id),
            "month": "2026-09",
            "product_status": str(product.status_id),
            "category": str(product.category_id),
            "product": [str(product.id)],
            "planning_activity": "ALL",
            "method": "INCREASE_PERCENT",
            "parameter": "10",
            "reason": "Target campaign September",
        }

        preview = self.client.post(reverse("sales:planning_builder"), payload)

        self.assertEqual(preview.status_code, 200)
        self.assertContains(preview, "PROJECTION PREVIEW")
        row = preview.context["builder_preview"]["rows"][0]
        self.assertEqual(row["product"], product)
        self.assertEqual([cell["month"] for cell in row["history"]], [date(2026, 6, 1), date(2026, 7, 1), date(2026, 8, 1)])
        self.assertEqual(row["history"][-1]["qty"], Decimal("62"))
        self.assertTrue(row["history"][-1]["is_projection"])
        self.assertEqual(row["target_gross"], Decimal("6800000"))
        self.assertEqual(row["target_qty"], 68)
        self.assertEqual(preview.context["builder_preview"]["parent_rows"][0]["parent_sku"], "PARENT-BUILDER")
        self.assertEqual(len(preview.context["builder_preview"]["sku_rows"]), 2)
        self.assertEqual(sum(cell["target_qty"] for cell in preview.context["builder_preview"]["sku_rows"]), 68)
        self.assertContains(preview, "Parent SKU (Sum)")
        self.assertContains(preview, 'data-preview-grain-panel="sku"')
        self.assertContains(preview, 'data-preview-grain-panel="parent_sku"')
        self.assertContains(preview, 'data-preview-grain-selector')
        self.assertContains(preview, 'name="preview_grain"', count=2)
        self.assertContains(preview, 'data-preview-table-scroll', count=2)
        self.assertContains(preview, "Cancel Preview")
        self.assertContains(preview, 'data-sales-planning-total', count=3)  # Two footers and the refresh selector.
        self.assertContains(preview, '<tfoot>', count=2)
        totals = preview.context["builder_preview"]["totals"]
        self.assertEqual(totals["qty"], 68)
        self.assertEqual(totals["gross"], Decimal("6800000"))
        self.assertEqual(totals["history"][-1]["qty"], Decimal("62"))

        payload["action"] = "save"
        payload[f"target_qty_{sku.id}"] = "15"
        for sku_row in preview.context["builder_preview"]["sku_rows"]:
            payload.setdefault(f"target_qty_{sku_row['sku'].id}", str(sku_row["target_qty"]))
        saved = self.client.post(reverse("sales:planning_builder"), payload, follow=True)
        plan = SalesPlan.objects.get(scenario=scenario, month=date(2026, 9, 1), product=product)
        self.assertEqual(saved.status_code, 200)
        self.assertEqual(plan.gross_sales_target, Decimal("1500000"))
        self.assertEqual(plan.quantity_target, 15)
        self.assertEqual(plan.sku_targets.get(sku=sku).quantity_target, 15)
        self.assertContains(saved, "Preview September 2026 tersimpan")
        self.assertContains(saved, 'name="draft_grain"')
        self.assertContains(saved, "Parent SKU (Sum)")
        self.assertContains(saved, "Close Draft")
        self.assertContains(saved, "Save Draft", count=2)
        self.assertContains(saved, 'id="sales-scenario-draft-form"')
        self.assertContains(saved, 'data-dirty-submit')
        self.assertContains(saved, 'data-dirty-submit-button', count=2)
        self.assertContains(saved, 'disabled aria-disabled="true"', count=2)
        self.assertContains(saved, 'form="sales-scenario-draft-form"')
        self.assertContains(saved, 'data-preview-table-scroll', count=2)
        self.assertContains(saved, 'data-draft-selection-delete-form')
        self.assertContains(saved, 'data-draft-row-select')
        self.assertContains(saved, 'title="Delete Selected SKU From Draft"')
        self.assertNotContains(saved, "Qty Gap")
        self.assertNotContains(saved, "Gross Gap")
        self.assertEqual(
            [header["month"] for header in saved.context["draft_history_headers"]],
            [date(2026, 6, 1), date(2026, 7, 1), date(2026, 8, 1)],
        )
        self.assertEqual(saved.context["rows"][0]["history"][-1]["qty"], Decimal("62"))
        self.assertEqual(saved.context["draft_parent_rows"][0]["parent_sku"], "PARENT-BUILDER")
        self.assertEqual(saved.context["draft_parent_rows"][0]["target_qty"], 15)
        self.assertContains(saved, '<tfoot>', count=3)
        self.assertEqual(saved.context["target_totals"]["qty"], 15)
        self.assertEqual(saved.context["target_totals"]["gross"], Decimal("1500000"))
        self.assertEqual(saved.context["target_totals"]["history"][-1]["qty"], Decimal("62"))

        closed = self.client.get(reverse("sales:planning_builder"), {
            "scenario": str(scenario.id),
            "month": "2026-09",
            "close_draft": "1",
        })
        self.assertEqual(closed.status_code, 200)
        self.assertNotContains(closed, 'id="draft-projection"')
        self.assertContains(closed, 'id="scenario-library"')

    def test_sales_builder_excludes_saved_products_only_for_the_selected_month(self):
        draft_product, draft_sku = self._planning_product("DRAFT-TARGET")
        approved_product, _ = self._planning_product("APPROVED-TARGET")
        available_product, _ = self._planning_product("AVAILABLE-TARGET")
        planned_subcategory = Subcategory.objects.create(
            category=draft_product.category, code="PLANNED", name="Planned subcategory",
        )
        available_subcategory = Subcategory.objects.create(
            category=draft_product.category, code="AVAILABLE", name="Available subcategory",
        )
        Product.objects.filter(pk__in=[draft_product.pk, approved_product.pk]).update(subcategory=planned_subcategory)
        Product.objects.filter(pk=available_product.pk).update(subcategory=available_subcategory)
        scenario = SalesPlanningScenario.objects.create(
            name="October targets", start_month=date(2026, 10, 1),
            end_month=date(2026, 11, 1), created_by=self.user,
        )
        approved = SalesPlanningScenario.objects.create(
            name="Other approved scenario", start_month=date(2026, 10, 1),
            end_month=date(2026, 10, 1), created_by=self.user,
            status=SalesPlanningScenario.Status.APPROVED,
        )
        plan = SalesPlan.objects.create(scenario=scenario, month=date(2026, 10, 1), product=draft_product)
        SalesPlanSKU.objects.create(plan=plan, sku=draft_sku, quantity_target=0)
        SalesPlan.objects.create(scenario=approved, month=date(2026, 10, 1), product=approved_product)
        filters = {
            "target_month": "2026-10", "planning_activity": "ALL",
            "product_status": str(draft_product.status_id), "category": str(draft_product.category_id),
        }
        options = self.client.get(reverse("sales:planning_filter_options"), filters).json()
        self.assertEqual([row["id"] for row in options["products"]], [str(available_product.id)])
        self.assertIn(str(draft_product.category_id), [row["id"] for row in options["categories"]])
        self.assertEqual([row["id"] for row in options["subcategories"]], [str(available_subcategory.id)])
        stale_subcategory = self.client.get(reverse("sales:planning_filter_options"), {
            **filters, "subcategory": str(planned_subcategory.id),
        }).json()
        self.assertEqual(stale_subcategory, options)
        next_month = self.client.get(reverse("sales:planning_filter_options"), {
            **filters, "target_month": "2026-11",
        }).json()
        self.assertEqual(len(next_month["products"]), 3)
        page = self.client.get(reverse("sales:planning_builder"), {"scenario": scenario.id, "month": "2026-10"})
        self.assertEqual(list(page.context["product_options"]), [available_product])
        preview = self.client.post(reverse("sales:planning_builder"), {
            "form_name": "builder", "action": "preview", "scenario": scenario.id,
            "month": "2026-10", "planning_activity": "ALL", "method": "SAME_AS_LAST_MONTH",
        })
        self.assertEqual({row["product"].id for row in preview.context["builder_preview"]["sku_rows"]}, {available_product.id})
        # The category disappears only when its last eligible product is planned.
        SalesPlan.objects.create(scenario=scenario, month=date(2026, 10, 1), product=available_product)
        completed = self.client.get(reverse("sales:planning_filter_options"), filters).json()
        self.assertEqual(completed, {"categories": [], "subcategories": [], "products": []})
        future = self.client.get(reverse("sales:planning_filter_options"), {
            **filters, "target_month": "2026-11",
        }).json()
        self.assertEqual([row["id"] for row in future["categories"]], [str(draft_product.category_id)])
        self.assertEqual(len(future["subcategories"]), 2)
        self.assertEqual(len(future["products"]), 3)

    def test_sales_builder_duplicate_submission_cannot_overwrite_existing_target(self):
        product, sku = self._planning_product("DUPLICATE-TARGET")
        scenario = SalesPlanningScenario.objects.create(
            name="Original target", start_month=date(2026, 10, 1),
            end_month=date(2026, 11, 1), created_by=self.user,
        )
        other_scenario = SalesPlanningScenario.objects.create(
            name="Different scenario", start_month=date(2026, 10, 1),
            end_month=date(2026, 11, 1), created_by=self.user,
        )
        payload = {
            "form_name": "builder", "action": "save", "scenario": str(scenario.id),
            "month": "2026-10", "product": str(product.id),
            "planning_activity": "ALL", "method": "SAME_AS_LAST_MONTH", f"target_qty_{sku.id}": "12",
        }
        self.assertEqual(self.client.post(reverse("sales:planning_builder"), payload).status_code, 302)
        target = SalesPlanSKU.objects.get(sku=sku)
        for scenario_id in (scenario.id, other_scenario.id):
            for action in ("preview", "save"):
                response = self.client.post(reverse("sales:planning_builder"), {
                    **payload, "scenario": scenario_id, "action": action, f"target_qty_{sku.id}": "99",
                }, follow=True)
                self.assertContains(response, "Product sudah memiliki Sales Projection pada bulan ini")
        target.refresh_from_db()
        self.assertEqual(target.quantity_target, 12)
        self.assertEqual(SalesPlan.objects.count(), 1)
        with self.assertRaises(IntegrityError), transaction.atomic():
            SalesPlan.objects.create(scenario=other_scenario, month=date(2026, 10, 1), product=product)
        # Revising the existing draft remains the supported edit path.
        self.client.post(reverse("sales:planning_builder"), {
            "form_name": "projection", "scenario": scenario.id, "month": "2026-10", f"qty_{target.id}": "15",
        })
        target.refresh_from_db()
        self.assertEqual(target.quantity_target, 15)
        self.client.post(reverse("sales:planning_builder"), {**payload, "month": "2026-11"})
        self.assertEqual(SalesPlan.objects.count(), 2)

    def test_sales_builder_rechecks_stale_preview_atomically(self):
        product, sku = self._planning_product("STALE-TARGET")
        new_product, new_sku = self._planning_product("STILL-AVAILABLE")
        scenario = SalesPlanningScenario.objects.create(
            name="Stale preview", start_month=date(2026, 10, 1),
            end_month=date(2026, 10, 1), created_by=self.user,
        )
        payload = {
            "form_name": "builder", "action": "preview", "scenario": scenario.id,
            "month": "2026-10", "planning_activity": "ALL", "method": "SAME_AS_LAST_MONTH",
        }
        preview = self.client.post(reverse("sales:planning_builder"), payload).context["builder_preview"]
        plan = SalesPlan.objects.create(scenario=scenario, month=date(2026, 10, 1), product=product, quantity_target=7)
        target = SalesPlanSKU.objects.create(plan=plan, sku=sku, quantity_target=7)
        save_payload = {**payload, "action": "save", f"target_qty_{sku.id}": "99", f"target_qty_{new_sku.id}": "10"}
        request = RequestFactory().post(reverse("sales:planning_builder"), save_payload)
        request.user = self.user
        with self.assertRaisesMessage(ValidationError, "Product sudah memiliki Sales Projection"):
            _save_sales_projection_preview(request, preview)
        response = self.client.post(reverse("sales:planning_builder"), save_payload, follow=True)
        self.assertContains(response, "Pilihan Product sudah berubah")
        self.assertFalse(SalesPlan.objects.filter(product=new_product).exists())
        target.refresh_from_db()
        self.assertEqual(target.quantity_target, 7)

    def test_sales_projection_parent_sku_view_sums_multiple_products(self):
        first, _ = self._planning_product("PARENT-SUM-A")
        second, _ = self._planning_product("PARENT-SUM-B")
        Product.objects.filter(id__in=[first.id, second.id]).update(parent_sku="PARENT-SUM")
        scenario = SalesPlanningScenario.objects.create(
            name="Parent Summary",
            start_month=date(2026, 9, 1),
            end_month=date(2026, 9, 1),
            created_by=self.user,
        )

        preview = self.client.post(reverse("sales:planning_builder"), {
            "form_name": "builder",
            "action": "preview",
            "scenario": str(scenario.id),
            "month": "2026-09",
            "product": [str(first.id), str(second.id)],
            "planning_activity": "ALL",
            "method": "SAME_AS_LAST_MONTH",
        })

        self.assertEqual(preview.status_code, 200)
        parent_rows = preview.context["builder_preview"]["parent_rows"]
        self.assertEqual(len(parent_rows), 1)
        self.assertEqual(parent_rows[0]["parent_sku"], "PARENT-SUM")
        self.assertEqual(parent_rows[0]["product_count"], 2)
        self.assertEqual(parent_rows[0]["sku_count"], 2)

    def test_sales_scenario_draft_deletes_selected_sku_or_parent_across_all_months(self):
        product, first_sku = self._planning_product("DELETE-DRAFT")
        product.parent_sku = "PARENT-DELETE"
        product.save(update_fields=["parent_sku"])
        second_sku = SKU.objects.create(
            sku="SKU-DELETE-DRAFT-L",
            product_variant=first_sku.product_variant,
            size="L",
            current_retail_price=Decimal("100000"),
        )
        scenario = SalesPlanningScenario.objects.create(
            name="Delete Draft Rows",
            start_month=date(2026, 9, 1),
            end_month=date(2026, 10, 1),
            created_by=self.user,
        )
        for month in (date(2026, 9, 1), date(2026, 10, 1)):
            plan = SalesPlan.objects.create(
                scenario=scenario,
                month=month,
                product=product,
                gross_sales_target=Decimal("3000000"),
                quantity_target=30,
            )
            SalesPlanSKU.objects.create(
                plan=plan,
                sku=first_sku,
                gross_sales_target=Decimal("1000000"),
                quantity_target=10,
            )
            SalesPlanSKU.objects.create(
                plan=plan,
                sku=second_sku,
                gross_sales_target=Decimal("2000000"),
                quantity_target=20,
            )

        deleted_sku = self.client.post(reverse("sales:planning_builder"), {
            "form_name": "delete_selection",
            "scenario": str(scenario.id),
            "month": "2026-09",
            "selection_grain": "sku",
            "selected_item": [str(first_sku.id)],
        }, follow=True)
        self.assertEqual(deleted_sku.status_code, 200)
        self.assertFalse(SalesPlanSKU.objects.filter(plan__scenario=scenario, sku=first_sku).exists())
        self.assertEqual(SalesPlanSKU.objects.filter(plan__scenario=scenario, sku=second_sku).count(), 2)
        self.assertEqual(
            set(SalesPlan.objects.filter(scenario=scenario).values_list("quantity_target", flat=True)),
            {20},
        )

        deleted_parent = self.client.post(reverse("sales:planning_builder"), {
            "form_name": "delete_selection",
            "scenario": str(scenario.id),
            "month": "2026-09",
            "selection_grain": "parent_sku",
            "selected_item": ["PARENT-DELETE"],
        }, follow=True)
        self.assertEqual(deleted_parent.status_code, 200)
        self.assertFalse(SalesPlanSKU.objects.filter(plan__scenario=scenario).exists())
        self.assertFalse(SalesPlan.objects.filter(scenario=scenario).exists())
        self.assertContains(deleted_parent, "berhasil dihapus dari seluruh bulan Scenario Draft")

    def test_transaction_export_uses_active_filters_and_exports_all_columns(self):
        for source, source_label, order_number, order_day in (
            (SalesOrder.Source.SHOPEE, "Shopee", "EXPORT-SHOPEE", date(2026, 8, 10)),
            (SalesOrder.Source.TIKTOK, "Tiktok", "EXPORT-TIKTOK", date(2026, 8, 11)),
        ):
            order = SalesOrder.objects.create(
                source=source,
                source_label=source_label,
                order_number=order_number,
                order_datetime=timezone.make_aware(datetime.combine(order_day, datetime.min.time())),
                order_date=order_day,
                current_status="Selesai",
                source_status="Selesai",
                is_final=True,
                first_seen_batch_id=uuid.uuid4(),
                latest_batch_id=uuid.uuid4(),
            )
            SalesOrderLine.objects.create(
                order=order,
                sku_code_snapshot=f"SKU-{source_label}",
                product_name_snapshot=f"Product {source_label}",
                variant_name_snapshot="M",
                quantity=2,
                net_unit_price=Decimal("90000"),
                retail_price_snapshot=Decimal("100000"),
                sales_cogs_snapshot=Decimal("50000"),
                total_gross_sales=Decimal("200000"),
                total_net_sales=Decimal("180000"),
                total_cogs=Decimal("100000"),
                gpm=Decimal("80000"),
                gpm_rate=Decimal("0.44444444"),
            )

        response = self.client.get(reverse("sales:transactions"), {
            "date_from": "2026-08-01",
            "date_to": "2026-08-31",
            "source": "Shopee",
            "q": "EXPORT",
            "export": "xlsx",
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("vobia-transactions_2026-08-01_2026-08-31.xlsx", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
        rows = list(workbook["Transactions"].iter_rows(values_only=True))
        self.assertEqual(rows[0][0:5], ("Order Date", "Order Datetime", "Shipped Datetime", "Source", "Order Number"))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][3], "Shopee")
        self.assertEqual(rows[1][4], "EXPORT-SHOPEE")
        self.assertEqual(rows[1][16], 2)
        self.assertEqual(rows[1][20], 180000)
        self.assertNotIn("EXPORT-TIKTOK", {cell for row in rows for cell in row})

    def test_report_filters_use_auto_apply_without_apply_buttons(self):
        for name in ("sales:dashboard", "sales:product_performance", "sales:pareto"):
            with self.subTest(name=name):
                response = self.client.get(reverse(name))
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, "data-auto-submit")
                self.assertNotContains(response, ">Apply Filters<")

    def test_pareto_supports_month_quarter_semester_and_year_periods(self):
        fixtures = (
            (date(2026, 1, 10), "PARETO-JAN", "100000"),
            (date(2026, 4, 10), "PARETO-APR", "200000"),
            (date(2026, 8, 10), "PARETO-AUG", "400000"),
        )
        for order_day, order_number, net in fixtures:
            order = SalesOrder.objects.create(
                source=SalesOrder.Source.SHOPEE,
                source_label="Shopee",
                order_number=order_number,
                order_datetime=timezone.make_aware(datetime.combine(order_day, datetime.min.time())),
                order_date=order_day,
                current_status="Selesai",
                source_status="Selesai",
                is_final=True,
                first_seen_batch_id=uuid.uuid4(),
                latest_batch_id=uuid.uuid4(),
            )
            net_value = Decimal(net)
            SalesOrderLine.objects.create(
                order=order,
                sku_code_snapshot=order_number,
                product_name_snapshot=order_number,
                quantity=1,
                net_unit_price=net_value,
                retail_price_snapshot=net_value,
                sales_cogs_snapshot=Decimal("50000"),
                total_gross_sales=net_value,
                total_net_sales=net_value,
                total_cogs=Decimal("50000"),
                gpm=net_value - Decimal("50000"),
            )

        default_response = self.client.get(reverse("sales:pareto"))
        self.assertEqual(default_response.context["period_type"], "year")
        self.assertEqual(default_response.context["period_value"], "2026")
        self.assertEqual(default_response.context["totals"]["net"], Decimal("700000"))
        self.assertContains(default_response, ">2026</option>")

        cases = (
            ("month", "2026-01", Decimal("100000")),
            ("quarter", "2026-Q2", Decimal("200000")),
            ("semester", "2026-S2", Decimal("400000")),
            ("year", "2026", Decimal("700000")),
        )
        for period_type, period, expected_net in cases:
            with self.subTest(period_type=period_type, period=period):
                response = self.client.get(reverse("sales:pareto"), {
                    "period_type": period_type,
                    "period": period,
                })
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.context["period_type"], period_type)
                self.assertEqual(response.context["period_value"], period)
                self.assertEqual(response.context["totals"]["net"], expected_net)

    def test_dashboard_custom_date_range_filters_all_sales_metrics(self):
        for day, order_number, net in ((5, "AUG-05", "100000"), (15, "AUG-15", "250000")):
            net_value = Decimal(net)
            gross_value = net_value + Decimal("10000")
            order = SalesOrder.objects.create(
                source=SalesOrder.Source.SHOPEE,
                source_label="Shopee",
                order_number=order_number,
                order_datetime=timezone.make_aware(datetime(2026, 8, day, 10, 0)),
                order_date=date(2026, 8, day),
                current_status="Selesai",
                source_status="Selesai",
                is_final=True,
                first_seen_batch_id=uuid.uuid4(),
                latest_batch_id=uuid.uuid4(),
            )
            SalesOrderLine.objects.create(
                order=order,
                sku_code_snapshot=f"SKU-{day}",
                product_name_snapshot="Test Product",
                quantity=1,
                net_unit_price=net_value,
                retail_price_snapshot=gross_value,
                sales_cogs_snapshot=Decimal("50000"),
                total_gross_sales=gross_value,
                total_net_sales=net_value,
                total_cogs=Decimal("50000"),
                gpm=net_value - Decimal("50000"),
            )

        response = self.client.get(reverse("sales:dashboard"), {
            "date_from": "2026-08-01",
            "date_to": "2026-08-10",
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["date_from"], date(2026, 8, 1))
        self.assertEqual(response.context["date_to"], date(2026, 8, 10))
        self.assertEqual(response.context["totals"]["orders"], 1)
        self.assertEqual(response.context["totals"]["net"], Decimal("100000"))
        self.assertEqual(response.context["source_rows"][0]["orders"], 1)
        self.assertEqual(response.context["period_type"], "custom")
        self.assertEqual(response.context["period_trend_grain"], "day")
        self.assertEqual(len(response.context["period_trend"]), 10)
        self.assertEqual(response.context["period_trend"][4]["day"], date(2026, 8, 5))
        self.assertEqual(response.context["period_trend"][4]["label"], "05 Agu")
        self.assertEqual(response.context["period_trend"][4]["net"], Decimal("100000"))
        self.assertEqual(response.context["period_trend"][4]["gross"], Decimal("110000"))
        self.assertEqual(response.context["period_trend"][5]["net"], 0)
        self.assertContains(response, "Daily Gross Sales")
        self.assertNotContains(response, "Daily Gross Sales dalam periode")
        self.assertContains(response, 'name="date_from"')
        self.assertContains(response, 'name="date_to"')
        self.assertContains(response, 'data-max-visible-rows="25"')

    def test_dashboard_supports_month_quarter_semester_year_and_trend_grain(self):
        for order_day, order_number, gross in (
            (date(2026, 1, 10), "DASH-JAN", "100000"),
            (date(2026, 4, 10), "DASH-APR", "200000"),
            (date(2026, 8, 10), "DASH-AUG", "400000"),
        ):
            gross_value = Decimal(gross)
            order = SalesOrder.objects.create(
                source=SalesOrder.Source.SHOPEE,
                source_label="Shopee",
                order_number=order_number,
                order_datetime=timezone.make_aware(datetime.combine(order_day, datetime.min.time())),
                order_date=order_day,
                current_status="Selesai",
                source_status="Selesai",
                is_final=True,
                first_seen_batch_id=uuid.uuid4(),
                latest_batch_id=uuid.uuid4(),
            )
            SalesOrderLine.objects.create(
                order=order,
                sku_code_snapshot=order_number,
                product_name_snapshot="Dashboard Period Product",
                quantity=1,
                net_unit_price=gross_value,
                retail_price_snapshot=gross_value,
                sales_cogs_snapshot=Decimal("0"),
                total_gross_sales=gross_value,
                total_net_sales=gross_value,
                total_cogs=Decimal("0"),
                gpm=gross_value,
            )

        cases = (
            ("month", "2026-08", "day", 31, Decimal("400000"), "Daily Gross Sales"),
            ("quarter", "2026-Q2", "month", 3, Decimal("200000"), "Monthly Gross Sales Trend"),
            ("semester", "2026-S2", "month", 6, Decimal("400000"), "Monthly Gross Sales Trend"),
            ("year", "2026", "month", 12, Decimal("700000"), "Monthly Gross Sales Trend"),
        )
        for period_type, period, grain, row_count, expected_gross, title in cases:
            with self.subTest(period_type=period_type, period=period):
                response = self.client.get(reverse("sales:dashboard"), {
                    "period_type": period_type,
                    "period": period,
                })
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.context["period_type"], period_type)
                self.assertEqual(response.context["period_value"], period)
                self.assertEqual(response.context["period_trend_grain"], grain)
                self.assertEqual(len(response.context["period_trend"]), row_count)
                self.assertEqual(response.context["totals"]["gross"], expected_gross)
                self.assertContains(response, title)

    def test_dashboard_monthly_gross_chart_uses_rolling_twelve_months(self):
        for order_day, order_number, gross in (
            (date(2026, 1, 10), "JAN-2026", "1000000000"),
            (date(2026, 3, 10), "MAR-2026", "2000000000"),
            (date(2027, 2, 10), "FEB-2027", "3000000000"),
        ):
            order = SalesOrder.objects.create(
                source=SalesOrder.Source.SHOPEE,
                source_label="Shopee",
                order_number=order_number,
                order_datetime=timezone.make_aware(datetime.combine(order_day, datetime.min.time())),
                order_date=order_day,
                current_status="Selesai",
                source_status="Selesai",
                is_final=True,
                first_seen_batch_id=uuid.uuid4(),
                latest_batch_id=uuid.uuid4(),
            )
            gross_value = Decimal(gross)
            SalesOrderLine.objects.create(
                order=order,
                sku_code_snapshot=order_number,
                product_name_snapshot="Test Product",
                quantity=1,
                net_unit_price=gross_value,
                retail_price_snapshot=gross_value,
                sales_cogs_snapshot=Decimal("0"),
                total_gross_sales=gross_value,
                total_net_sales=gross_value,
                total_cogs=Decimal("0"),
                gpm=gross_value,
            )

        response = self.client.get(reverse("sales:dashboard"))
        monthly = response.context["monthly_gross"]

        self.assertEqual(len(monthly), 12)
        self.assertEqual(monthly[0]["month"], date(2026, 3, 1))
        self.assertEqual(monthly[-1]["month"], date(2027, 2, 1))
        self.assertEqual(monthly[0]["gross_billion"], Decimal("2"))
        self.assertEqual(monthly[-1]["gross_billion"], Decimal("3"))
        self.assertEqual(response.context["mtd_cutoff_day"], 10)
        self.assertEqual(response.context["mtd_gross"][0]["gross_billion"], Decimal("2"))
        self.assertEqual(response.context["mtd_gross"][-1]["gross_billion"], Decimal("3"))
        self.assertNotContains(response, "Jan: Rp 1.000.000.000")
        self.assertContains(response, 'class="panel monthly-sales-panel"')
        self.assertContains(response, "MTD Gross Sales")
        self.assertContains(response, 'class="panel trend-panel"')
        self.assertLess(response.content.index(b"Daily Gross Sales"), response.content.index(b"Monthly Gross Sales"))
        self.assertLess(response.content.index(b"Monthly Gross Sales"), response.content.index(b"MTD Gross Sales"))
        self.assertLess(response.content.index(b"MTD Gross Sales"), response.content.index(b"Status periode"))

    def test_dashboard_source_group_and_source_are_cascading_multi_filters(self):
        fixtures = (
            (SalesOrder.Source.SHOPEE, "Shopee", "100000"),
            (SalesOrder.Source.TIKTOK, "Tiktok", "200000"),
            (SalesOrder.Source.OTHER, "Whatsapp", "400000"),
        )
        for index, (source, source_label, gross) in enumerate(fixtures, start=1):
            order = SalesOrder.objects.create(
                source=source,
                source_label=source_label,
                order_number=f"DASH-SOURCE-{index}",
                order_datetime=timezone.make_aware(datetime(2026, 8, index, 10, 0)),
                order_date=date(2026, 8, index),
                current_status="Selesai",
                source_status="Selesai",
                is_final=True,
                first_seen_batch_id=uuid.uuid4(),
                latest_batch_id=uuid.uuid4(),
            )
            gross_value = Decimal(gross)
            SalesOrderLine.objects.create(
                order=order,
                sku_code_snapshot=f"DASH-SOURCE-SKU-{index}",
                product_name_snapshot="Dashboard Source Product",
                quantity=1,
                net_unit_price=gross_value,
                retail_price_snapshot=gross_value,
                sales_cogs_snapshot=Decimal("0"),
                total_gross_sales=gross_value,
                total_net_sales=gross_value,
                total_cogs=Decimal("0"),
                gpm=gross_value,
            )

        response = self.client.get(reverse("sales:dashboard"), {
            "date_from": "2026-08-01",
            "date_to": "2026-08-31",
            "source_group": ["Marketplace"],
            "source": ["Shopee", "Tiktok"],
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_source_groups"], ["Marketplace"])
        self.assertEqual(response.context["selected_sources"], ["Shopee", "Tiktok"])
        self.assertEqual(response.context["totals"]["orders"], 2)
        self.assertEqual(response.context["totals"]["gross"], Decimal("300000"))
        self.assertEqual(response.context["monthly_gross"][-1]["gross"], Decimal("300000"))
        self.assertEqual(
            {item["value"] for item in response.context["source_options"]},
            {"Shopee", "Tiktok"},
        )
        self.assertEqual(response.content.count(b"data-multi-select data-all-label"), 2)
        self.assertContains(response, 'data-source-group="Marketplace"')
        self.assertContains(response, "data-source-cascade")

        stale_source = self.client.get(reverse("sales:dashboard"), {
            "date_from": "2026-08-01",
            "date_to": "2026-08-31",
            "source_group": ["Marketplace"],
            "source": ["Whatsapp"],
        })
        self.assertEqual(stale_source.context["selected_sources"], [])
        self.assertEqual(stale_source.context["totals"]["orders"], 2)

        all_sources = self.client.get(reverse("sales:dashboard"), {
            "date_from": "2026-08-01",
            "date_to": "2026-08-31",
        })
        self.assertEqual(all_sources.context["selected_source_groups"], [])
        self.assertEqual(all_sources.context["selected_sources"], [])
        self.assertEqual(all_sources.context["totals"]["orders"], 3)
        self.assertEqual(all_sources.context["totals"]["gross"], Decimal("700000"))

    def test_dashboard_mtd_gross_uses_latest_sales_day_as_same_monthly_cutoff(self):
        for order_day, order_number, gross in (
            (date(2026, 1, 19), "JAN-19", "1000000000"),
            (date(2026, 1, 20), "JAN-20", "2000000000"),
            (date(2026, 2, 19), "FEB-19", "6000000000"),
            (date(2026, 3, 19), "MAR-19", "3000000000"),
        ):
            order = SalesOrder.objects.create(
                source=SalesOrder.Source.SHOPEE,
                source_label="Shopee",
                order_number=order_number,
                order_datetime=timezone.make_aware(datetime.combine(order_day, datetime.min.time())),
                order_date=order_day,
                current_status="Selesai",
                source_status="Selesai",
                is_final=True,
                first_seen_batch_id=uuid.uuid4(),
                latest_batch_id=uuid.uuid4(),
            )
            gross_value = Decimal(gross)
            SalesOrderLine.objects.create(
                order=order,
                sku_code_snapshot=order_number,
                product_name_snapshot="Test Product",
                quantity=1,
                net_unit_price=gross_value,
                retail_price_snapshot=gross_value,
                sales_cogs_snapshot=Decimal("0"),
                total_gross_sales=gross_value,
                total_net_sales=gross_value,
                total_cogs=Decimal("0"),
                gpm=gross_value,
            )

        response = self.client.get(reverse("sales:dashboard"))
        monthly = {row["month"]: row for row in response.context["monthly_gross"]}
        mtd = {row["month"]: row for row in response.context["mtd_gross"]}

        self.assertEqual(response.context["mtd_cutoff_day"], 19)
        self.assertEqual(monthly[date(2026, 1, 1)]["gross"], Decimal("3000000000"))
        self.assertEqual(mtd[date(2026, 1, 1)]["gross"], Decimal("1000000000"))
        self.assertEqual(mtd[date(2026, 2, 1)]["gross"], Decimal("6000000000"))
        self.assertEqual(mtd[date(2026, 2, 1)]["growth_pct"], Decimal("500"))
        self.assertEqual(mtd[date(2026, 2, 1)]["growth_class"], "positive")
        self.assertEqual(mtd[date(2026, 3, 1)]["growth_pct"], Decimal("-50.0"))
        self.assertEqual(mtd[date(2026, 3, 1)]["growth_class"], "negative")
        self.assertEqual(monthly[date(2026, 2, 1)]["growth_pct"], Decimal("100"))
        self.assertEqual(monthly[date(2026, 2, 1)]["growth_class"], "positive")
        self.assertEqual(monthly[date(2026, 3, 1)]["growth_pct"], Decimal("-50.0"))
        self.assertEqual(monthly[date(2026, 3, 1)]["growth_class"], "negative")
        self.assertContains(response, "Tanggal 1–19 setiap bulan")
        self.assertContains(response, 'class="monthly-growth-divider"')
        self.assertContains(response, 'class="growth-positive"')
        self.assertContains(response, 'class="growth-negative"')
        self.assertContains(response, 'aria-label="Growth MTD Gross Sales bulanan"')

    def test_product_performance_accepts_multiple_product_filters(self):
        for index, (product_name, net) in enumerate((
            ("Product Alpha", "100000"),
            ("Product Beta", "200000"),
            ("Product Gamma", "400000"),
        ), start=1):
            order = SalesOrder.objects.create(
                source=SalesOrder.Source.SHOPEE,
                source_label="Shopee",
                order_number=f"MULTI-{index}",
                order_datetime=timezone.make_aware(datetime(2026, 8, index, 10, 0)),
                order_date=date(2026, 8, index),
                current_status="Selesai",
                source_status="Selesai",
                is_final=True,
                first_seen_batch_id=uuid.uuid4(),
                latest_batch_id=uuid.uuid4(),
            )
            net_value = Decimal(net)
            SalesOrderLine.objects.create(
                order=order,
                sku_code_snapshot=f"MULTI-SKU-{index}",
                product_name_snapshot=product_name,
                quantity=1,
                net_unit_price=net_value,
                retail_price_snapshot=net_value,
                sales_cogs_snapshot=Decimal("50000"),
                total_gross_sales=net_value,
                total_net_sales=net_value,
                total_cogs=Decimal("50000"),
                gpm=net_value - Decimal("50000"),
            )

        response = self.client.get(reverse("sales:product_performance"), {
            "date_from": "2026-08-01",
            "date_to": "2026-08-31",
            "product": ["Product Alpha", "Product Beta"],
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_products"], ["Product Alpha", "Product Beta"])
        self.assertEqual(response.context["totals"]["orders"], 2)
        self.assertEqual(response.context["totals"]["net"], Decimal("300000"))
        self.assertContains(response, "2 selected")
        self.assertContains(response, 'data-filter-search')
        self.assertContains(response, 'data-filter-select-visible')

    def test_product_performance_links_product_to_parent_traffic_listing(self):
        status = ProductStatus.objects.create(code="TRAFFIC", name="Regular")
        category = Category.objects.create(code="PANTS", name="Pants")
        product = Product.objects.create(code="PARENT-1", parent_sku="PARENT-1", name="Denim Product", status=status, category=category)
        second_product = Product.objects.create(code="PARENT-2", parent_sku="PARENT-2", name="Denim Product 2", status=status, category=category)
        variant = ProductVariant.objects.create(product=product, name="Blue")
        second_variant = ProductVariant.objects.create(product=second_product, name="Black")
        sku = SKU.objects.create(sku="DENIM.M", product_variant=variant, size="M")
        second_sku = SKU.objects.create(sku="DENIM2.M", product_variant=second_variant, size="M")
        MarketplaceProductMapping.objects.create(source="Tiktok", marketplace_product_code="TIKTOK-PARENT", product=product)
        MarketplaceProductMapping.objects.create(source="Tiktok", marketplace_product_code="TIKTOK-PARENT", product=second_product)
        order = SalesOrder.objects.create(
            source=SalesOrder.Source.TIKTOK,
            source_label="Tiktok",
            order_number="TRAFFIC-LINK-1",
            order_datetime=timezone.make_aware(datetime(2026, 8, 10, 10, 0)),
            order_date=date(2026, 8, 10),
            current_status="Selesai",
            source_status="Selesai",
            is_final=True,
            first_seen_batch_id=uuid.uuid4(),
            latest_batch_id=uuid.uuid4(),
        )
        SalesOrderLine.objects.create(
            order=order,
            sku=sku,
            product_name_snapshot=product.name,
            quantity=1,
            net_unit_price=Decimal("100000"),
            total_gross_sales=Decimal("100000"),
            total_net_sales=Decimal("100000"),
        )
        SalesOrderLine.objects.create(
            order=order,
            sku=second_sku,
            product_name_snapshot=second_product.name,
            quantity=1,
            net_unit_price=Decimal("100000"),
            total_gross_sales=Decimal("100000"),
            total_net_sales=Decimal("100000"),
        )
        raw = RawFile.objects.create(
            dataset_type=RawFile.DatasetType.TRAFFIC_TIKTOK,
            original_filename="traffic.csv",
            storage_path="tests/traffic.csv",
            checksum_sha256="a" * 64,
            byte_size=1,
            detected_format="csv",
            uploaded_by=self.user,
        )
        batch = TrafficImportBatch.objects.create(
            raw_file=raw,
            source="Tiktok",
            period_start=date(2026, 8, 1),
            period_end=date(2026, 8, 25),
            status=TrafficImportBatch.Status.COMMITTED,
        )
        for key, name in (("LISTING", "Parent Traffic Listing"), ("PRODUCT", product.name)):
            TrafficProductMetric.objects.create(
                source="Tiktok",
                period_start=date(2026, 8, 1),
                period_end=date(2026, 8, 25),
                product=product if key == "PRODUCT" else None,
                traffic_product_key=key,
                marketplace_product_code_snapshot="TIKTOK-PARENT",
                product_name_snapshot=name,
                views=100,
                clicks=20,
                visitors=10,
                source_batch=batch,
            )

        by_product = self.client.get(reverse("sales:product_performance"), {
            "date_from": "2026-08-01",
            "date_to": "2026-08-25",
            "source": "Tiktok",
            "product": [product.name, second_product.name],
        })
        self.assertEqual(by_product.context["selected_products"], [product.name, second_product.name])
        self.assertEqual(by_product.context["traffic_totals"], {"views": 100, "clicks": 20, "visitors": 10})

    def test_product_performance_accepts_multiple_product_status_filters(self):
        for index, (product_status, net) in enumerate((
            ("Regular", "100000"),
            ("Essential+", "200000"),
            ("Discontinue", "400000"),
        ), start=1):
            order = SalesOrder.objects.create(
                source=SalesOrder.Source.SHOPEE,
                source_label="Shopee",
                order_number=f"MULTI-STATUS-{index}",
                order_datetime=timezone.make_aware(datetime(2026, 8, index, 10, 0)),
                order_date=date(2026, 8, index),
                current_status="Selesai",
                source_status="Selesai",
                is_final=True,
                first_seen_batch_id=uuid.uuid4(),
                latest_batch_id=uuid.uuid4(),
            )
            net_value = Decimal(net)
            SalesOrderLine.objects.create(
                order=order,
                sku_code_snapshot=f"MULTI-STATUS-SKU-{index}",
                product_name_snapshot=f"Product Status {index}",
                product_status_snapshot=product_status,
                quantity=1,
                net_unit_price=net_value,
                retail_price_snapshot=net_value,
                sales_cogs_snapshot=Decimal("50000"),
                total_gross_sales=net_value,
                total_net_sales=net_value,
                total_cogs=Decimal("50000"),
                gpm=net_value - Decimal("50000"),
            )

        response = self.client.get(reverse("sales:product_performance"), {
            "date_from": "2026-08-01",
            "date_to": "2026-08-31",
            "product_status": ["Regular", "Essential+"],
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_product_statuses"], ["Regular", "Essential+"])
        self.assertEqual(response.context["totals"]["orders"], 2)
        self.assertEqual(response.context["totals"]["net"], Decimal("300000"))
        self.assertContains(response, "Status Produk")
        self.assertContains(response, "2 selected")

    def test_product_performance_accepts_multiple_source_group_and_category_filters(self):
        fixtures = (
            (SalesOrder.Source.SHOPEE, "Shopee", "Category Alpha", "100000"),
            (SalesOrder.Source.TIKTOK, "Tiktok", "Category Beta", "200000"),
            (SalesOrder.Source.OTHER, "Whatsapp", "Category Gamma", "400000"),
        )
        for index, (source, source_label, category, net) in enumerate(fixtures, start=1):
            order = SalesOrder.objects.create(
                source=source,
                source_label=source_label,
                order_number=f"MULTI-DIM-{index}",
                order_datetime=timezone.make_aware(datetime(2026, 8, index, 10, 0)),
                order_date=date(2026, 8, index),
                current_status="Selesai",
                source_status="Selesai",
                is_final=True,
                first_seen_batch_id=uuid.uuid4(),
                latest_batch_id=uuid.uuid4(),
            )
            net_value = Decimal(net)
            SalesOrderLine.objects.create(
                order=order,
                sku_code_snapshot=f"MULTI-DIM-SKU-{index}",
                product_name_snapshot=f"Dimension Product {index}",
                category_snapshot=category,
                quantity=1,
                net_unit_price=net_value,
                retail_price_snapshot=net_value,
                sales_cogs_snapshot=Decimal("50000"),
                total_gross_sales=net_value,
                total_net_sales=net_value,
                total_cogs=Decimal("50000"),
                gpm=net_value - Decimal("50000"),
            )

        response = self.client.get(reverse("sales:product_performance"), {
            "date_from": "2026-08-01",
            "date_to": "2026-08-31",
            "source": ["Shopee", "Tiktok"],
            "source_group": ["Marketplace"],
            "category": ["Category Alpha", "Category Beta"],
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_sources"], ["Shopee", "Tiktok"])
        self.assertEqual(response.context["selected_source_groups"], ["Marketplace"])
        self.assertEqual(response.context["selected_categories"], ["Category Alpha", "Category Beta"])
        self.assertEqual(response.context["totals"]["orders"], 2)
        self.assertEqual(response.context["totals"]["net"], Decimal("300000"))
        self.assertEqual(response.content.count(b"data-multi-select data-all-label"), 5)

    def test_product_performance_cascades_status_category_and_product_options(self):
        fixtures = (
            ("Regular", "Knitwear", "Product Knit"),
            ("Regular", "Shirt", "Product Shirt"),
            ("Seasonal New", "Socks", "Product Socks"),
        )
        for index, (product_status, category, product_name) in enumerate(fixtures, start=1):
            order = SalesOrder.objects.create(
                source=SalesOrder.Source.SHOPEE,
                source_label="Shopee",
                order_number=f"CASCADE-{index}",
                order_datetime=timezone.make_aware(datetime(2026, 8, index, 10, 0)),
                order_date=date(2026, 8, index),
                current_status="Selesai",
                source_status="Selesai",
                is_final=True,
                first_seen_batch_id=uuid.uuid4(),
                latest_batch_id=uuid.uuid4(),
            )
            SalesOrderLine.objects.create(
                order=order,
                sku_code_snapshot=f"CASCADE-SKU-{index}",
                product_name_snapshot=product_name,
                product_status_snapshot=product_status,
                category_snapshot=category,
                quantity=1,
                net_unit_price=Decimal("100000"),
                retail_price_snapshot=Decimal("100000"),
                sales_cogs_snapshot=Decimal("50000"),
                total_gross_sales=Decimal("100000"),
                total_net_sales=Decimal("100000"),
                total_cogs=Decimal("50000"),
                gpm=Decimal("50000"),
            )

        stale = self.client.get(
            reverse("sales:product_performance"),
            {
                "date_from": "2026-08-01",
                "date_to": "2026-08-31",
                "product_status": ["Regular"],
                "category": ["Socks"],
                "product": ["Product Socks"],
            },
        )
        self.assertEqual(stale.status_code, 200)
        self.assertEqual(stale.context["categories"], ("Knitwear", "Shirt"))
        self.assertEqual(stale.context["selected_categories"], [])
        self.assertEqual(stale.context["products"], ("Product Knit", "Product Shirt"))
        self.assertEqual(stale.context["selected_products"], [])
        self.assertEqual(stale.context["totals"]["orders"], 2)

        valid = self.client.get(
            reverse("sales:product_performance"),
            {
                "date_from": "2026-08-01",
                "date_to": "2026-08-31",
                "product_status": ["Regular"],
                "category": ["Knitwear"],
            },
        )
        self.assertEqual(valid.status_code, 200)
        self.assertEqual(valid.context["products"], ("Product Knit",))
        self.assertEqual(valid.context["selected_categories"], ["Knitwear"])
        self.assertEqual(valid.context["totals"]["orders"], 1)

from .services.manual import create_manual_sale
from .services.requirements import import_requirements, summarize_import_requirements


class ManualSalesAndRequirementTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="sales", password="test-password")
        status = ProductStatus.objects.create(code="ACTIVE", name="Active")
        category = Category.objects.create(code="APPAREL", name="Apparel")
        product = Product.objects.create(code="P-1", name="Product", status=status, category=category)
        variant = ProductVariant.objects.create(product=product, name="Black")
        self.sku = SKU.objects.create(
            sku="SKU-1",
            product_variant=variant,
            current_retail_price=Decimal("200000"),
            current_master_cogs=Decimal("100000"),
        )

    def test_manual_sale_preserves_actual_source_and_posts_fifo_exception(self):
        line = create_manual_sale(
            source_label="Whatsapp",
            order_number="WA-001",
            order_datetime=timezone.make_aware(datetime(2026, 8, 19, 10, 0)),
            sku=self.sku,
            quantity=2,
            net_unit_price=Decimal("200000"),
            status="Selesai",
            actor=self.user,
        )
        self.assertEqual(line.business_key, "Whatsapp|WA-001|SKU-1")
        self.assertEqual(line.retail_price_snapshot, Decimal("200000"))
        self.assertEqual(line.total_gross_sales, line.total_net_sales)
        movement = InventoryMovement.objects.get(sales_line=line)
        self.assertEqual(movement.movement_key, "SALES|Whatsapp|WA-001|SKU-1")
        self.assertEqual(movement.allocated_cost, Decimal("0"))
        self.assertEqual(InventoryException.objects.get(movement=movement).quantity, Decimal("2"))

    def test_manual_sale_adjusts_only_transaction_snapshot_and_marks_special_case(self):
        line = create_manual_sale(
            source_label="Whatsapp",
            order_number="WA-ABOVE-RETAIL",
            order_datetime=timezone.make_aware(datetime(2026, 8, 19, 10, 0)),
            sku=self.sku,
            quantity=1,
            net_unit_price=Decimal("220000"),
            status="Selesai",
            actor=self.user,
        )
        self.sku.refresh_from_db()
        self.assertEqual(self.sku.current_retail_price, Decimal("200000"))
        self.assertEqual(line.retail_price_snapshot, Decimal("220000"))
        self.assertTrue(line.retail_price_special_case)
        self.assertEqual(line.total_gross_sales, line.total_net_sales)

    def test_manual_sale_special_case_shows_warning_notification(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sales:input_transaction"),
            {
                "source_label": "Whatsapp",
                "order_number": "WA-NOTIFICATION",
                "order_datetime": "2026-08-19T10:00",
                "sku": self.sku.pk,
                "quantity": 1,
                "net_unit_price": "220000",
                "status": "Selesai",
                "shipped": "",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "SPECIAL CASE HARGA")
        self.assertContains(response, "Retail Price master tetap Rp 200.000")

    def test_manual_transaction_is_separate_from_data_import(self):
        self.client.force_login(self.user)

        manual_response = self.client.get(reverse("sales:input_transaction"))
        import_response = self.client.get(reverse("imports:sales_list"))

        self.assertContains(manual_response, "Input Transaction")
        self.assertContains(manual_response, "Post transaksi")
        self.assertNotContains(import_response, "Input transaksi manual")
        self.assertNotContains(import_response, "Post transaksi")

    def test_nonfinal_old_month_remains_in_import_requirement(self):
        SalesOrder.objects.create(
            source=SalesOrder.Source.SHOPEE,
            source_label="Shopee",
            order_number="OLD-001",
            order_datetime=timezone.make_aware(datetime(2026, 4, 10, 10, 0)),
            order_date=datetime(2026, 4, 10).date(),
            current_status="Sedang Dikirim",
            source_status="Sedang Dikirim",
            is_final=False,
            first_seen_batch_id="11111111-1111-1111-1111-111111111111",
            latest_batch_id="11111111-1111-1111-1111-111111111111",
        )
        requirements = import_requirements(as_of_date=datetime(2026, 8, 20).date())
        april = next(row for row in requirements if row["source"] == "Shopee" and row["period_start"].month == 4)
        self.assertEqual(april["period_start"], date(2026, 4, 10))
        self.assertEqual(april["nonfinal_count"], 1)
        self.assertIn("Sedang Dikirim", april["nonfinal_statuses"])

        summary = summarize_import_requirements(requirements, as_of_date=date(2026, 8, 20))
        shopee_summary = next(row for row in summary if row["source"] == SalesOrder.Source.SHOPEE)
        self.assertEqual(
            shopee_summary,
            {
                "source": SalesOrder.Source.SHOPEE,
                "period_start": date(2026, 4, 10),
                "period_end": date(2026, 8, 19),
            },
        )

    def test_requirement_summary_collapses_months_into_one_range_per_marketplace(self):
        requirements = [
            {
                "source": SalesOrder.Source.TIKTOK,
                "period_start": date(2026, 7, 1),
                "period_end": date(2026, 7, 31),
            },
            {
                "source": SalesOrder.Source.SHOPEE,
                "period_start": date(2026, 8, 1),
                "period_end": date(2026, 8, 19),
            },
            {
                "source": SalesOrder.Source.TIKTOK,
                "period_start": date(2026, 6, 1),
                "period_end": date(2026, 8, 19),
            },
        ]

        self.assertEqual(
            summarize_import_requirements(requirements, as_of_date=date(2026, 8, 20)),
            [
                {
                    "source": SalesOrder.Source.SHOPEE,
                    "period_start": date(2026, 8, 1),
                    "period_end": date(2026, 8, 19),
                },
                {
                    "source": SalesOrder.Source.TIKTOK,
                    "period_start": date(2026, 6, 1),
                    "period_end": date(2026, 8, 19),
                },
            ],
        )
