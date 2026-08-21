import csv
import math
import uuid
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from audit.services import record_audit
from imports.models import RawFile
from imports.services.storage import DuplicateRawFile, _checksum
from master_data.models import MarketplaceProductMapping

from ..models import (
    StagedTrafficRow,
    TrafficImportBatch,
    TrafficImportIssue,
    TrafficPeriodState,
    TrafficProductMetric,
)


ALIASES = {
    "code": {"id produk", "kode produk", "product id", "kode produk shopee", "kode produk tiktok"},
    "name": {"produk", "nama produk", "product", "product name"},
    "views": {"jumlah produk dilihat", "dilihat", "product views", "impresi produk", "product impressions"},
    "clicks": {"produk diklik", "klik produk", "product clicks", "clicks"},
    "visitors": {"pengunjung produk", "pengunjung", "kunjungan", "klik unik", "unique clicks", "unique visitors"},
}


def _norm(value):
    return " ".join(str(value or "").strip().lower().split())


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        if not value.is_integer():
            raise ValidationError("Kode produk terbaca sebagai angka desimal; ekspor ulang kolom ID sebagai text.")
        return str(int(value))
    return str(value).strip()


def _count(value):
    if value in (None, ""):
        return 0
    text = str(value).strip().replace(".", "").replace(",", "")
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValidationError(f"Metric traffic bukan angka: {value}") from exc
    if number < 0 or number != number.to_integral_value():
        raise ValidationError(f"Metric traffic harus bilangan bulat nonnegative: {value}")
    return int(number)


def _read(path, detected_format):
    if detected_format == "csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(8192)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
            rows = list(csv.reader(handle, dialect=dialect))
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            rows = [list(row) for row in workbook.active.iter_rows(values_only=True)]
        finally:
            workbook.close()
    if not rows:
        raise ValidationError("File traffic kosong.")
    headers = [_text(value) for value in rows[0]]
    data = []
    for row_number, values in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in values):
            continue
        padded = list(values) + [None] * max(0, len(headers) - len(values))
        data.append((row_number, dict(zip(headers, padded, strict=False))))
    return headers, data


@transaction.atomic
def parse_batch(batch):
    batch.staged_rows.all().delete()
    batch.issues.all().delete()
    path = Path(settings.PRIVATE_UPLOAD_ROOT) / batch.raw_file.storage_path
    try:
        headers, rows = _read(path, batch.raw_file.detected_format)
    except ValidationError as exc:
        TrafficImportIssue.objects.create(batch=batch, severity="ERROR", code="FILE_READ_ERROR", message=" ".join(exc.messages), is_blocking=True)
        batch.status = TrafficImportBatch.Status.BLOCKED
        batch.blocking_issue_count = 1
        batch.save(update_fields=["status", "blocking_issue_count"])
        return batch
    normalized = {_norm(header): header for header in headers}
    resolved = {}
    for logical, aliases in ALIASES.items():
        matches = [normalized[alias] for alias in aliases if alias in normalized]
        if len(matches) == 1:
            resolved[logical] = matches[0]
        elif not matches:
            TrafficImportIssue.objects.create(batch=batch, severity="ERROR", code="MISSING_HEADER", field_name=logical, message=f"Header logis '{logical}' tidak ditemukan. Header tersedia: {', '.join(headers)}", is_blocking=True)
        else:
            TrafficImportIssue.objects.create(batch=batch, severity="ERROR", code="AMBIGUOUS_HEADER", field_name=logical, message=f"Lebih dari satu header cocok untuk '{logical}': {', '.join(matches)}", is_blocking=True)
    if batch.issues.filter(is_blocking=True).exists():
        batch.status = TrafficImportBatch.Status.BLOCKED
        batch.total_rows = len(rows)
        batch.blocking_issue_count = batch.issues.filter(is_blocking=True).count()
        batch.quality_summary = {"available_headers": headers, "resolved_headers": resolved}
        batch.save()
        return batch
    staged = []
    code_counts = Counter()
    for row_number, source in rows:
        try:
            code = _text(source.get(resolved["code"]))
            views = _count(source.get(resolved["views"]))
            clicks = _count(source.get(resolved["clicks"]))
            visitors = _count(source.get(resolved["visitors"]))
        except ValidationError as exc:
            TrafficImportIssue.objects.create(batch=batch, severity="ERROR", code="INVALID_ROW", field_name="metrics", message=f"Row {row_number}: {' '.join(exc.messages)}", is_blocking=True)
            continue
        if not code:
            TrafficImportIssue.objects.create(batch=batch, severity="ERROR", code="MISSING_PRODUCT_CODE", field_name="code", message=f"Row {row_number}: kode produk kosong.", is_blocking=True)
            continue
        mappings = MarketplaceProductMapping.objects.filter(source=batch.source, marketplace_product_code=code, is_active=True).select_related("product")
        product = mappings.first().product if mappings.count() == 1 else None
        row = StagedTrafficRow.objects.create(
            batch=batch,
            row_number=row_number,
            marketplace_product_code=code,
            product_name=_text(source.get(resolved["name"])),
            product=product,
            views=views,
            clicks=clicks,
            visitors=visitors,
            selected_source_data={resolved[key]: source.get(resolved[key]) for key in resolved},
        )
        staged.append(row)
        code_counts[code] += 1
        if mappings.count() != 1:
            TrafficImportIssue.objects.create(batch=batch, staged_row=row, severity="ERROR", code="PRODUCT_MAPPING_NOT_UNIQUE", field_name="code", message="Kode marketplace harus memetakan tepat ke satu Product canonical.", is_blocking=True)
    for row in staged:
        if code_counts[row.marketplace_product_code] > 1:
            TrafficImportIssue.objects.create(batch=batch, staged_row=row, severity="ERROR", code="DUPLICATE_PRODUCT_CODE", field_name="code", message="Kode produk muncul lebih dari sekali. Pilih/ekspor hanya baris utama product agar traffic variasi tidak terduplikasi.", is_blocking=True)
    batch.total_rows = len(rows)
    batch.ready_rows = len(staged)
    batch.blocking_issue_count = batch.issues.filter(is_blocking=True).count()
    batch.warning_count = batch.issues.filter(severity="WARNING").count()
    batch.status = TrafficImportBatch.Status.BLOCKED if batch.blocking_issue_count else TrafficImportBatch.Status.READY
    batch.quality_summary = {"available_headers": headers, "resolved_headers": resolved, "unique_product_codes": len(code_counts)}
    batch.save()
    return batch


def create_traffic_import(uploaded_file, source, period_start, period_end, actor):
    if period_start > period_end or period_start.year != period_end.year or period_start.month != period_end.month:
        raise ValidationError("Satu batch traffic harus berada dalam satu bulan yang sama.")
    if period_start.day != 1:
        raise ValidationError("Traffic bulanan/MTD harus dimulai dari tanggal pertama bulan.")
    month = period_start.replace(day=1)
    state = TrafficPeriodState.objects.filter(source=source, month=month).first()
    if state and state.is_complete:
        raise ValidationError("Periode traffic sudah complete. Re-open terlebih dahulu agar perubahan tercatat.")
    extension = Path(uploaded_file.name).suffix.lower().lstrip(".")
    checksum = _checksum(uploaded_file)
    dataset_type = RawFile.DatasetType.TRAFFIC_SHOPEE if source == TrafficPeriodState.Source.SHOPEE else RawFile.DatasetType.TRAFFIC_TIKTOK
    duplicate = RawFile.objects.filter(dataset_type=dataset_type, checksum_sha256=checksum).first()
    if duplicate:
        raise DuplicateRawFile(duplicate)
    relative = Path("traffic") / source.lower() / str(period_start.year) / f"{period_start.month:02d}" / f"{uuid.uuid4()}.{extension}"
    absolute = Path(settings.PRIVATE_UPLOAD_ROOT) / relative
    absolute.parent.mkdir(parents=True, exist_ok=True)
    with absolute.open("wb") as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)
    uploaded_file.seek(0)
    try:
        with transaction.atomic():
            raw = RawFile.objects.create(
                dataset_type=dataset_type,
                original_filename=Path(uploaded_file.name).name,
                storage_path=str(relative),
                checksum_sha256=checksum,
                byte_size=uploaded_file.size,
                detected_format=extension,
                uploaded_by=actor,
                source_metadata={"source": source, "dataset": "product_traffic", "period_start": str(period_start), "period_end": str(period_end)},
            )
            batch = TrafficImportBatch.objects.create(raw_file=raw, source=source, period_start=period_start, period_end=period_end)
            record_audit(actor=actor, action="traffic_import_uploaded", entity_type="traffic.trafficimportbatch", entity_id=batch.id, metadata={"source": source, "checksum_sha256": checksum})
    except Exception:
        absolute.unlink(missing_ok=True)
        raise
    return parse_batch(batch)


@transaction.atomic
def commit_batch(batch_id, actor):
    batch = TrafficImportBatch.objects.select_for_update().get(pk=batch_id)
    if batch.status != TrafficImportBatch.Status.READY or batch.issues.filter(is_blocking=True).exists():
        raise ValidationError("Batch traffic belum siap atau masih memiliki blocking issue.")
    for row in batch.staged_rows.select_related("product"):
        TrafficProductMetric.objects.update_or_create(
            source=batch.source,
            period_start=batch.period_start,
            traffic_product_key=row.marketplace_product_code,
            defaults={
                "period_end": batch.period_end,
                "product": row.product,
                "marketplace_product_code_snapshot": row.marketplace_product_code,
                "product_name_snapshot": row.product_name,
                "category_snapshot": row.product.category.name if row.product_id else "",
                "subcategory_snapshot": row.product.subcategory.name if row.product_id and row.product.subcategory_id else "",
                "is_historical_migration": False,
                "views": row.views,
                "clicks": row.clicks,
                "visitors": row.visitors,
                "source_batch": batch,
            },
        )
    now = timezone.now()
    state, _ = TrafficPeriodState.objects.get_or_create(source=batch.source, month=batch.period_start.replace(day=1))
    state.last_successful_import_at = now
    state.last_data_end = batch.period_end
    state.save()
    batch.status = TrafficImportBatch.Status.COMMITTED
    batch.approved_by = actor
    batch.approved_at = now
    batch.save(update_fields=["status", "approved_by", "approved_at"])
    record_audit(actor=actor, action="traffic_import_committed", entity_type="traffic.trafficimportbatch", entity_id=batch.id, after_values={"records": batch.staged_rows.count(), "period": str(batch.period_start)})
    return batch


@transaction.atomic
def mark_period_complete(source, month, actor):
    month = month.replace(day=1)
    if month >= timezone.localdate().replace(day=1):
        raise ValidationError("Bulan berjalan tidak boleh ditandai complete.")
    state = TrafficPeriodState.objects.select_for_update().filter(source=source, month=month).first()
    if not state or not state.last_successful_import_at:
        raise ValidationError("Periode belum memiliki successful import.")
    state.is_complete = True
    state.save(update_fields=["is_complete"])
    record_audit(actor=actor, action="traffic_period_completed", entity_type="traffic.trafficperiodstate", entity_id=state.id, after_values={"source": source, "month": str(month)})
    return state


@transaction.atomic
def reopen_period(source, month, actor, reason):
    if not reason.strip():
        raise ValidationError("Alasan re-open wajib diisi.")
    state = TrafficPeriodState.objects.select_for_update().get(source=source, month=month.replace(day=1))
    if not state.is_complete:
        raise ValidationError("Periode belum berstatus complete.")
    state.is_complete = False
    state.reopen_count += 1
    state.save(update_fields=["is_complete", "reopen_count"])
    record_audit(actor=actor, action="traffic_period_reopened", entity_type="traffic.trafficperiodstate", entity_id=state.id, reason=reason)
    return state
