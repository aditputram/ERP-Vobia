import io

from django.contrib.auth.decorators import login_required
from django.db.models import Count, Max, Min, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from imports.services.master_parser import CANONICAL_HEADERS

from .models import (
    Category,
    MarketplaceProductMapping,
    Product,
    SKU,
    Supplier,
    Warehouse,
)


def _active_product_codes():
    mappings = {}
    for mapping in MarketplaceProductMapping.objects.filter(is_active=True).order_by(
        "product_id", "source", "-created_at"
    ):
        mappings.setdefault((mapping.product_id, mapping.source), mapping.marketplace_product_code)
    return mappings


@login_required
def overview(request):
    query = request.GET.get("q", "").strip()
    grain = request.GET.get("grain", "sku")
    if grain not in {"sku", "parent"}:
        grain = "sku"
    mappings = _active_product_codes()

    if grain == "parent":
        rows = Product.objects.select_related("status", "category", "subcategory").annotate(
            sku_count=Count("variants__skus", distinct=True),
            cogs_min=Min("variants__skus__current_master_cogs"),
            cogs_max=Max("variants__skus__current_master_cogs"),
            retail_min=Min("variants__skus__current_retail_price"),
            retail_max=Max("variants__skus__current_retail_price"),
        )
        if query:
            rows = rows.filter(
                Q(name__icontains=query)
                | Q(article__icontains=query)
                | Q(parent_sku__icontains=query)
                | Q(code__icontains=query)
                | Q(variants__skus__sku__icontains=query)
                | Q(marketplace_mappings__marketplace_product_code__icontains=query)
            ).distinct()
    else:
        rows = SKU.objects.select_related(
            "product_variant__product__status",
            "product_variant__product__category",
            "product_variant__product__subcategory",
        )
        if query:
            rows = rows.filter(
                Q(sku__icontains=query)
                | Q(product_variant__product__name__icontains=query)
                | Q(product_variant__product__article__icontains=query)
                | Q(product_variant__product__parent_sku__icontains=query)
                | Q(product_variant__product__code__icontains=query)
                | Q(
                    product_variant__product__marketplace_mappings__marketplace_product_code__icontains=query
                )
            ).distinct()

    rows = list(rows)
    for row in rows:
        product = row if grain == "parent" else row.product_variant.product
        row.shopee_code = mappings.get(
            (product.id, MarketplaceProductMapping.Source.SHOPEE), ""
        )
        row.tiktok_code = mappings.get(
            (product.id, MarketplaceProductMapping.Source.TIKTOK), ""
        )

    sku_quality = {
        "missing_cogs": SKU.objects.filter(current_master_cogs__isnull=True).count(),
        "missing_retail": SKU.objects.filter(current_retail_price__isnull=True).count(),
        "missing_both": SKU.objects.filter(
            current_master_cogs__isnull=True,
            current_retail_price__isnull=True,
        ).count(),
        "inactive": SKU.objects.filter(is_active=False).count(),
    }
    context = {
        "counts": {
            "products": Product.objects.count(),
            "skus": SKU.objects.count(),
            "categories": Category.objects.count(),
            "suppliers": Supplier.objects.count(),
            "warehouses": Warehouse.objects.count(),
        },
        "sku_quality": sku_quality,
        "rows": rows,
        "grain": grain,
        "query": query,
    }
    return render(request, "master_data/overview.html", context)


@login_required
def export_bank_data(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "VOBIA"
    sheet.append(CANONICAL_HEADERS)
    sheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="111111")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    mappings = _active_product_codes()

    skus = SKU.objects.select_related(
        "product_variant__product__status",
        "product_variant__product__category",
        "product_variant__product__subcategory",
    ).order_by("sku")
    for sku in skus:
        variant = sku.product_variant
        product = variant.product
        sheet.append(
            [
                "Vobia",
                sku.sku,
                product.parent_sku,
                product.article,
                product.category.name,
                product.subcategory.name if product.subcategory else "",
                "" if variant.name == "Default" else variant.name,
                sku.size,
                product.status.name,
                sku.current_master_cogs if sku.current_master_cogs is not None else "",
                sku.current_retail_price if sku.current_retail_price is not None else "",
                mappings.get((product.id, MarketplaceProductMapping.Source.SHOPEE), ""),
                mappings.get((product.id, MarketplaceProductMapping.Source.TIKTOK), ""),
            ]
        )

    for row in sheet.iter_rows(min_row=2):
        for column in (*range(1, 10), 12, 13):
            row[column - 1].number_format = "@"
        row[9].number_format = "#,##0.0000"
        row[10].number_format = "#,##0.00"

    widths = (12, 20, 20, 28, 18, 20, 18, 18, 20, 16, 16, 22, 22)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.auto_filter.ref = sheet.dimensions

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="VOBIA-Bank-Data-{timezone.localdate():%Y-%m-%d}.xlsx"'
    )
    return response
