import hashlib
from calendar import month_name
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from accounts.models import User
from master_data.models import SKU
from merchandising.models import MerchandisingMonthlySnapshot, MerchandisingSnapshotBatch


WORKBOOK_ID = "1crjvZPKrSSj2MFrysQZ3PWH5rUHrkhULzArjykFSlvU"
METRICS = {
    "incoming_qty": "Incoming QTY",
    "incoming_cogs": "Incoming COGS",
    "incoming_gross": "Incoming Gross",
    "beginning_qty": "Beginning QTY",
    "beginning_cogs": "Beginning COGS",
    "beginning_gross": "Beginning Gross",
    "sales_qty": "Sales QTY",
    "sales_cogs": "Sales COGS",
    "sales_gross": "Sales Gross",
    "sales_net": "Sales Net",
    "ratio": "Ratio",
    "ending_qty": "Ending Stock QTY",
    "ending_cogs": "Ending Stock COGS",
    "ending_gross": "Ending Stock Gross",
}


def clean_header(value):
    return " ".join(str(value or "").split())


def decimal_value(value, *, row, header):
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise CommandError(f"Nilai numerik tidak valid di row {row}, kolom {header}: {value!r}") from exc


class Command(BaseCommand):
    help = "Import immutable MD Actual baseline from a read-only Vobia MD 2026 XLSX export."

    def add_arguments(self, parser):
        parser.add_argument("file")
        parser.add_argument("--actor", default="vobiasuperadmin")

    def handle(self, *args, **options):
        source = Path(options["file"]).expanduser().resolve()
        if not source.is_file():
            raise CommandError(f"File tidak ditemukan: {source}")
        actor = User.objects.filter(username=options["actor"]).first()
        if actor is None:
            raise CommandError(f"User tidak ditemukan: {options['actor']}")

        checksum = hashlib.sha256(source.read_bytes()).hexdigest()
        existing = MerchandisingSnapshotBatch.objects.filter(source_sha256=checksum).first()
        if existing:
            self.stdout.write(self.style.WARNING(f"Snapshot identik sudah ada: {existing.id}"))
            return

        workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
        if "MD Actual" not in workbook.sheetnames:
            raise CommandError("Sheet MD Actual tidak ditemukan.")
        sheet = workbook["MD Actual"]
        iterator = sheet.iter_rows(min_row=4, max_row=1000, min_col=1, max_col=181, values_only=True)
        headers = [clean_header(value) for value in next(iterator)]
        header_index = {header: index for index, header in enumerate(headers) if header}

        required_identity = [
            "Status Product", "SKU", "ARTICLE", "Variant", "Category", "Sub Category",
            "Sub Variant", "COGS", "Retail Price", "December 25 Ending Stock QTY",
            "December 25 Ending Stock COGS", "December 25 Ending Stock Gross",
        ]
        missing_headers = [header for header in required_identity if header not in header_index]
        for month_number in range(1, 13):
            month = month_name[month_number]
            for label in METRICS.values():
                if f"{month} {label}" not in header_index:
                    missing_headers.append(f"{month} {label}")
        if missing_headers:
            raise CommandError(f"Header MD Actual belum sesuai: {', '.join(missing_headers[:10])}")

        source_rows = []
        source_skus = []
        for excel_row, row in enumerate(iterator, start=5):
            sku_code = str(row[header_index["SKU"]] or "").strip()
            if not sku_code:
                continue
            source_rows.append((excel_row, row, sku_code))
            source_skus.append(sku_code)
        if len(source_skus) != len(set(source_skus)):
            raise CommandError("Duplicate SKU ditemukan pada MD Actual.")

        master = SKU.objects.filter(sku__in=source_skus).in_bulk(field_name="sku")
        missing_skus = sorted(set(source_skus) - set(master))
        if missing_skus:
            raise CommandError(f"{len(missing_skus)} SKU MD Actual belum ada di master ERP: {', '.join(missing_skus[:10])}")

        source_mtime = timezone.make_aware(datetime.fromtimestamp(source.stat().st_mtime))
        snapshots = []
        with transaction.atomic():
            MerchandisingSnapshotBatch.objects.update(is_active=False)
            batch = MerchandisingSnapshotBatch.objects.create(
                source_workbook_id=WORKBOOK_ID,
                source_file_name=source.name,
                source_sha256=checksum,
                source_modified_at=source_mtime,
                imported_by=actor,
                row_count=len(source_rows),
                is_active=True,
            )
            for excel_row, row, sku_code in source_rows:
                identity = {
                    "status_snapshot": str(row[header_index["Status Product"]] or "").strip(),
                    "product_snapshot": str(row[header_index["ARTICLE"]] or "").strip(),
                    "variant_snapshot": str(row[header_index["Variant"]] or "").strip(),
                    "category_snapshot": str(row[header_index["Category"]] or "").strip(),
                    "subcategory_snapshot": str(row[header_index["Sub Category"]] or "").strip(),
                    "size_snapshot": str(row[header_index["Sub Variant"]] or "").strip(),
                    "cogs_snapshot": decimal_value(row[header_index["COGS"]], row=excel_row, header="COGS"),
                    "retail_price_snapshot": decimal_value(row[header_index["Retail Price"]], row=excel_row, header="Retail Price"),
                }
                prior = {
                    "prior_year_ending_qty": decimal_value(row[header_index["December 25 Ending Stock QTY"]], row=excel_row, header="December 25 Ending Stock QTY"),
                    "prior_year_ending_cogs": decimal_value(row[header_index["December 25 Ending Stock COGS"]], row=excel_row, header="December 25 Ending Stock COGS"),
                    "prior_year_ending_gross": decimal_value(row[header_index["December 25 Ending Stock Gross"]], row=excel_row, header="December 25 Ending Stock Gross"),
                }
                for month_number in range(1, 13):
                    month = month_name[month_number]
                    values = {
                        field: decimal_value(
                            row[header_index[f"{month} {label}"]],
                            row=excel_row,
                            header=f"{month} {label}",
                        )
                        for field, label in METRICS.items()
                    }
                    mos_header = f"{month} MOS"
                    mos = None
                    if mos_header in header_index and row[header_index[mos_header]] not in (None, ""):
                        mos = decimal_value(row[header_index[mos_header]], row=excel_row, header=mos_header)
                    snapshots.append(
                        MerchandisingMonthlySnapshot(
                            batch=batch,
                            sku=master[sku_code],
                            source_row=excel_row,
                            month=date(2026, month_number, 1),
                            mos=mos,
                            **identity,
                            **prior,
                            **values,
                        )
                    )
            MerchandisingMonthlySnapshot.objects.bulk_create(snapshots, batch_size=1000)

        self.stdout.write(
            self.style.SUCCESS(
                f"Imported {len(source_rows)} SKU dan {len(snapshots)} monthly snapshot. Batch: {batch.id}"
            )
        )
