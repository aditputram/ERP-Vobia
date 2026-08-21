import csv
import math
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from master_data.models import MarketplaceSKUMapping, SKU
from sales.models import SalesOrder, SalesOrderLine

from ..models import SalesImportBatch, SalesImportIssue, StagedSalesRow


PARSER_VERSION = "sales-v3"
SALES_CUTOVER_DATE = date(2026, 8, 1)
HEADER_ALIASES = {
    "Shopee": {
        "order_number": ("No. Pesanan",),
        "status": ("Status Pesanan",),
        "sku": ("Nomor Referensi SKU",),
        "quantity": ("Jumlah",),
        "net_unit_price": ("Harga Setelah Diskon",),
        "created_time": ("Waktu Pesanan Dibuat",),
        "shipped_time": ("Waktu Pengiriman Diatur",),
    },
    "Tiktok": {
        "order_number": ("Order ID",),
        "status": ("Order Status",),
        "sku": ("Seller SKU", "Seller SKU Input"),
        "quantity": ("Quantity", "SKU Sold Qty", "SKU sold quantity"),
        "original_unit_price": ("SKU Unit Original Price", "1 SKU original price"),
        "seller_discount": ("SKU Seller Discount", "Total seller discount in this SKU ID"),
        "created_time": ("Created Time",),
        "shipped_time": ("Shipped Time",),
    },
}


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal(value):
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        raise InvalidOperation
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            raise InvalidOperation
        return Decimal(str(value))
    text = str(value).strip().replace("Rp", "").replace("rp", "").replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        tail = text.rsplit(",", maxsplit=1)[1]
        text = text.replace(",", "" if len(tail) == 3 else ".")
    elif "." in text:
        tail = text.rsplit(".", maxsplit=1)[1]
        if len(tail) == 3 and text.replace(".", "").replace("-", "").isdigit():
            text = text.replace(".", "")
    return Decimal(text)


def _integer(value):
    parsed = _decimal(value)
    if parsed is None or parsed != parsed.to_integral_value():
        raise InvalidOperation
    return int(parsed)


def _datetime(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value).strip()
        parsed = None
        for format_string in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
        ):
            try:
                parsed = datetime.strptime(text, format_string)
                break
            except ValueError:
                continue
        if parsed is None:
            try:
                parsed = datetime.fromisoformat(text)
            except ValueError as exc:
                raise ValueError(f"Datetime tidak dikenali: {text}") from exc
    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def _resolve_headers(headers, source):
    resolved = {}
    missing = []
    for logical_name, aliases in HEADER_ALIASES[source].items():
        match = next((alias for alias in aliases if alias in headers), None)
        if match:
            resolved[logical_name] = match
        else:
            missing.append(" / ".join(aliases))
    return resolved, missing


def _read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(handle, dialect=dialect)
        headers = [_text(value) for value in next(reader)]
        rows = []
        for row_number, values in enumerate(reader, start=2):
            padded = list(values) + [None] * max(0, len(headers) - len(values))
            if not any(_text(value) for value in padded):
                continue
            rows.append((row_number, dict(zip(headers, padded, strict=False))))
        return headers, rows


def _read_xlsx(path, source):
    read_only = source == SalesImportBatch.Source.SHOPEE
    workbook = load_workbook(path, read_only=read_only, data_only=True)
    try:
        sheet = workbook.active
        if read_only:
            sheet.reset_dimensions()
        rows_iterator = sheet.iter_rows(values_only=True)
        headers = [_text(value) for value in next(rows_iterator)]
        rows = []
        for row_number, values in enumerate(rows_iterator, start=2):
            padded = list(values) + [None] * max(0, len(headers) - len(values))
            row = dict(zip(headers, padded, strict=False))
            if source == SalesImportBatch.Source.TIKTOK and row_number == 2:
                if _text(row.get("Order ID")).lower().startswith("platform unique order id"):
                    continue
            if not any(_text(value) for value in padded):
                continue
            rows.append((row_number, row))
        return headers, rows
    finally:
        workbook.close()


def _read_rows(path, file_format, source):
    if file_format == "csv":
        return _read_csv(path)
    if file_format == "xlsx":
        return _read_xlsx(path, source)
    raise ValueError("Format Sales import harus .xlsx atau .csv.")


def _normalize_status(source, source_status, shipped_datetime):
    status = source_status.strip()
    lowered = status.casefold()
    cancelled = lowered in {"batal", "dibatalkan", "cancelled", "canceled"}
    if cancelled:
        if shipped_datetime:
            return "Retur", True, False
        return "Batal", True, True
    if lowered == "selesai":
        return "Selesai", True, False
    return status, False, False


def _issue(batch, row, severity, code, field_name, message, blocking):
    return SalesImportIssue(
        batch=batch,
        staged_row=row,
        severity=severity,
        code=code,
        field_name=field_name,
        message=message,
        is_blocking=blocking,
    )


@transaction.atomic
def parse_sales_batch(batch):
    batch = SalesImportBatch.objects.select_for_update().select_related("raw_file").get(pk=batch.pk)
    if batch.status == SalesImportBatch.Status.COMMITTED:
        raise ValueError("Batch Sales committed tidak boleh diparsing ulang.")
    batch.staged_rows.all().delete()
    batch.issues.all().delete()
    path = Path(settings.PRIVATE_UPLOAD_ROOT) / batch.raw_file.storage_path
    try:
        headers, source_rows = _read_rows(path, batch.raw_file.detected_format, batch.source)
    except Exception as exc:
        SalesImportIssue.objects.create(
            batch=batch,
            severity=SalesImportIssue.Severity.ERROR,
            code="FILE_PARSE_ERROR",
            message=f"File tidak dapat diparsing: {exc}",
            is_blocking=True,
        )
        batch.status = SalesImportBatch.Status.BLOCKED
        batch.blocking_issue_count = 1
        batch.previewed_at = timezone.now()
        batch.save(update_fields=["status", "blocking_issue_count", "previewed_at"])
        return batch

    resolved, missing_headers = _resolve_headers(headers, batch.source)
    if missing_headers:
        SalesImportIssue.objects.create(
            batch=batch,
            severity=SalesImportIssue.Severity.ERROR,
            code="MISSING_SOURCE_HEADERS",
            field_name="header",
            message="Header wajib tidak ditemukan: " + ", ".join(missing_headers),
            is_blocking=True,
        )
        batch.status = SalesImportBatch.Status.BLOCKED
        batch.blocking_issue_count = 1
        batch.previewed_at = timezone.now()
        batch.quality_summary = {"headers": headers}
        batch.save()
        return batch

    if len(source_rows) > settings.SALES_IMPORT_MAX_ROWS:
        SalesImportIssue.objects.create(
            batch=batch,
            severity=SalesImportIssue.Severity.ERROR,
            code="ROW_LIMIT_EXCEEDED",
            message=f"File memiliki {len(source_rows)} baris; batas {settings.SALES_IMPORT_MAX_ROWS}.",
            is_blocking=True,
        )
        batch.status = SalesImportBatch.Status.BLOCKED
        batch.total_rows = len(source_rows)
        batch.blocking_issue_count = 1
        batch.previewed_at = timezone.now()
        batch.save()
        return batch

    marketplace_sku_mappings = {}
    if batch.source == SalesImportBatch.Source.TIKTOK:
        marketplace_sku_ids = {
            _text(row.get("SKU ID")) for _, row in source_rows if _text(row.get("SKU ID"))
        }
        marketplace_sku_mappings = {
            mapping.marketplace_sku_id: mapping
            for mapping in MarketplaceSKUMapping.objects.select_related("sku").filter(
                source=MarketplaceSKUMapping.Source.TIKTOK,
                marketplace_sku_id__in=marketplace_sku_ids,
                is_active=True,
            )
        }
    sku_values = [_text(row.get(resolved["sku"])) for _, row in source_rows]
    sku_values.extend(mapping.sku.sku for mapping in marketplace_sku_mappings.values())
    master_skus = {
        item.sku: item
        for item in SKU.objects.select_related("product_variant__product").filter(sku__in=sku_values)
    }
    order_numbers = [_text(row.get(resolved["order_number"])) for _, row in source_rows]
    existing_lines = {
        (line.order.order_number, line.sku.sku): line
        for line in SalesOrderLine.objects.select_related("order", "sku").filter(
            order__source=batch.source,
            order__order_number__in=order_numbers,
            sku__sku__in=sku_values,
        )
    }
    historical_orders = {
        order.order_number: order
        for order in SalesOrder.objects.filter(
            source=batch.source,
            order_number__in=order_numbers,
            import_origin=SalesOrder.ImportOrigin.HISTORICAL,
            order_date__lt=SALES_CUTOVER_DATE,
        )
    }
    historical_fallback_lines = {}
    if historical_orders:
        for line in SalesOrderLine.objects.select_related("order", "sku").filter(
            order_id__in=[order.id for order in historical_orders.values()]
        ).order_by("order_id", "committed_at"):
            historical_fallback_lines.setdefault(line.order.order_number, line)

    staged_rows = []
    pending_issues = []
    quality = Counter()
    dates = []
    out_of_scope_ids = set()
    historical_audit_order_numbers = set()

    for row_number, raw in source_rows:
        order_number = _text(raw.get(resolved["order_number"]))
        source_status = _text(raw.get(resolved["status"]))
        source_seller_sku = _text(raw.get(resolved["sku"]))
        marketplace_sku_id = _text(raw.get("SKU ID")) if batch.source == SalesImportBatch.Source.TIKTOK else ""
        sku_mapping = marketplace_sku_mappings.get(marketplace_sku_id)
        resolved_from_mapping = not source_seller_sku and sku_mapping is not None
        sku_text = sku_mapping.sku.sku if resolved_from_mapping else source_seller_sku
        row_issues = []
        quantity = None
        net_unit_price = None
        order_datetime = None
        shipped_datetime = None

        try:
            quantity = _integer(raw.get(resolved["quantity"]))
        except (InvalidOperation, ValueError):
            row_issues.append(("ERROR", "INVALID_QUANTITY", "quantity", "Qty harus berupa bilangan bulat positif.", True))
        try:
            order_datetime = _datetime(raw.get(resolved["created_time"]))
        except ValueError as exc:
            row_issues.append(("ERROR", "INVALID_ORDER_DATETIME", "created_time", str(exc), True))
        try:
            shipped_datetime = _datetime(raw.get(resolved["shipped_time"]))
        except ValueError as exc:
            row_issues.append(("ERROR", "INVALID_SHIPPED_DATETIME", "shipped_time", str(exc), True))

        if batch.source == SalesImportBatch.Source.SHOPEE:
            try:
                net_unit_price = _decimal(raw.get(resolved["net_unit_price"]))
            except (InvalidOperation, ValueError):
                row_issues.append(("ERROR", "INVALID_NET_PRICE", "net_unit_price", "Harga Setelah Diskon tidak valid.", True))
        else:
            try:
                original_price = _decimal(raw.get(resolved["original_unit_price"]))
                seller_discount = _decimal(raw.get(resolved["seller_discount"])) or Decimal("0")
                if quantity and quantity > 0 and original_price is not None:
                    net_unit_price = original_price - (seller_discount / quantity)
            except (InvalidOperation, ValueError, ZeroDivisionError):
                row_issues.append(("ERROR", "INVALID_NET_PRICE", "net_unit_price", "Harga/diskon TikTok tidak valid.", True))

        normalized_status, is_final, is_pure_cancelled = _normalize_status(
            batch.source,
            source_status,
            shipped_datetime,
        )
        is_out_of_scope = bool(
            order_datetime
            and timezone.localtime(order_datetime).date() < SALES_CUTOVER_DATE
        )
        historical_order = historical_orders.get(order_number) if is_out_of_scope else None
        historical_status_audit = historical_order is not None
        historical_status_update_allowed = historical_status_audit
        if is_out_of_scope:
            # Financial fields from pre-cutover raw exports never rewrite the
            # immutable historical transaction snapshot. A matched historical
            # order may only refresh its current status after approval.
            row_issues = []
        historical_line = None
        if historical_status_audit:
            historical_line = existing_lines.get((order_number, sku_text)) or historical_fallback_lines.get(order_number)
            quality["historical_status_audit_rows"] += 1
            historical_audit_order_numbers.add(order_number)
            if not source_status:
                row_issues.append(
                    (
                        "WARNING",
                        "HISTORICAL_STATUS_MISSING_IGNORED",
                        "status",
                        "Status raw historical kosong; status canonical tidak diubah.",
                        False,
                    )
                )
                normalized_status = historical_order.current_status
                is_final = historical_order.is_final
                is_pure_cancelled = historical_order.is_pure_cancelled
                historical_status_update_allowed = False
                quality["historical_status_ignored_rows"] += 1
            elif is_pure_cancelled:
                row_issues.append(
                    (
                        "WARNING",
                        "HISTORICAL_PURE_CANCELLATION_REVIEW",
                        "status",
                        "Pembatalan murni historical tidak menulis ulang transaksi/financial snapshot; "
                        "status canonical dipertahankan untuk review koreksi terpisah.",
                        False,
                    )
                )
                normalized_status = historical_order.current_status
                is_final = historical_order.is_final
                is_pure_cancelled = historical_order.is_pure_cancelled
                historical_status_update_allowed = False
                quality["historical_status_ignored_rows"] += 1
            elif (
                historical_order.current_status == "Retur" and normalized_status != "Retur"
            ) or (
                historical_order.is_final and not is_final
            ):
                row_issues.append(
                    (
                        "WARNING",
                        "HISTORICAL_FINAL_STATUS_REGRESSION_IGNORED",
                        "status",
                        f"Status final {historical_order.current_status} tidak boleh turun menjadi {source_status}; "
                        "status canonical tidak diubah.",
                        False,
                    )
                )
                normalized_status = historical_order.current_status
                is_final = historical_order.is_final
                is_pure_cancelled = historical_order.is_pure_cancelled
                historical_status_update_allowed = False
                quality["historical_status_ignored_rows"] += 1
        master_sku = historical_line.sku if historical_line else (None if is_out_of_scope else master_skus.get(sku_text))
        existing_line = historical_line if historical_status_audit else (None if is_out_of_scope else existing_lines.get((order_number, sku_text)))
        retail_snapshot = None
        master_retail_snapshot = None
        retail_price_special_case = False
        cogs_snapshot = None

        if not is_out_of_scope:
            for field_name, value in {
                "order_number": order_number,
                "status": source_status,
                "sku": sku_text,
            }.items():
                if not value:
                    row_issues.append(("ERROR", "REQUIRED_VALUE_MISSING", field_name, f"{field_name} wajib diisi.", True))
            if quantity is not None and quantity <= 0:
                row_issues.append(("ERROR", "NONPOSITIVE_QUANTITY", "quantity", "Qty harus lebih besar dari nol.", True))
            if net_unit_price is not None and net_unit_price < 0:
                row_issues.append(("ERROR", "NEGATIVE_NET_PRICE", "net_unit_price", "Harga net per unit tidak boleh negatif.", True))
            if master_sku is None and sku_text:
                row_issues.append(("ERROR", "UNKNOWN_SKU", "sku", f"SKU {sku_text} belum ada di Master Product ERP.", True))
            elif master_sku is not None:
                if master_sku.current_retail_price is None:
                    row_issues.append(("ERROR", "MISSING_RETAIL_PRICE", "retail_price", "Retail Price master kosong.", True))
                if master_sku.current_master_cogs is None:
                    row_issues.append(("ERROR", "MISSING_COGS", "cogs", "COGS master kosong.", True))
                if existing_line:
                    retail_snapshot = existing_line.retail_price_snapshot
                    cogs_snapshot = existing_line.sales_cogs_snapshot
                else:
                    master_retail_snapshot = master_sku.current_retail_price
                    retail_snapshot = master_retail_snapshot
                    cogs_snapshot = master_sku.current_master_cogs
                if (
                    existing_line is None
                    and retail_snapshot is not None
                    and net_unit_price is not None
                    and net_unit_price > retail_snapshot
                ):
                    retail_price_special_case = True
                    retail_snapshot = net_unit_price
                    quality["retail_price_special_case_rows"] += 1
                    row_issues.append(
                        (
                            "WARNING",
                            "RETAIL_PRICE_SNAPSHOT_SPECIAL_CASE",
                            "net_unit_price",
                            "Harga net per unit lebih besar dari Retail Price master. "
                            "Retail Price master tidak diubah; snapshot Retail Price hanya untuk transaksi ini "
                            "disamakan dengan harga net agar discount tidak negatif.",
                            False,
                        )
                    )

        if order_datetime:
            dates.append(order_datetime)
        if is_out_of_scope:
            quality["pre_cutover_rows"] += 1
        elif not is_final and not is_pure_cancelled:
            quality["nonfinal_rows"] += 1
            quality[f"status::{normalized_status}"] += 1
        if resolved_from_mapping and not is_out_of_scope:
            quality["resolved_marketplace_sku_rows"] += 1
        if is_pure_cancelled and not is_out_of_scope:
            quality["pure_cancel_rows"] += 1
        if normalized_status == "Retur" and not is_out_of_scope:
            quality["return_rows"] += 1

        selected_data = {
            "source_status": source_status,
            "order_number": order_number,
            "sku": sku_text,
            "source_seller_sku": source_seller_sku,
            "marketplace_sku_id": marketplace_sku_id,
            "sku_resolution": "marketplace_sku_mapping" if resolved_from_mapping else "seller_sku",
            "quantity": _text(raw.get(resolved["quantity"])),
            "created_time": _text(raw.get(resolved["created_time"])),
            "shipped_time": _text(raw.get(resolved["shipped_time"])),
            "import_scope": "before_cutover" if is_out_of_scope else "in_scope",
            "historical_status_audit": historical_status_audit,
            "historical_status_update_allowed": historical_status_update_allowed,
            "historical_order_id": str(historical_order.id) if historical_order else "",
            "financial_snapshot_locked": is_out_of_scope,
            "retail_price_rule": "transaction_special_case" if retail_price_special_case else "master_snapshot",
            "master_retail_price": str(master_retail_snapshot) if master_retail_snapshot is not None else "",
        }
        staged = StagedSalesRow.objects.create(
            batch=batch,
            row_number=row_number,
            order_number=order_number,
            source_label=batch.source,
            source_group="Marketplace",
            source_status=source_status,
            normalized_status=normalized_status,
            is_final=is_final,
            is_pure_cancelled=is_pure_cancelled,
            order_datetime=order_datetime,
            shipped_datetime=shipped_datetime,
            sku_text=sku_text,
            sku=master_sku,
            quantity=quantity,
            net_unit_price=net_unit_price,
            retail_price_snapshot=retail_snapshot,
            sales_cogs_snapshot=cogs_snapshot,
            product_status_snapshot=(master_sku.product_variant.product.status.name if master_sku else ""),
            category_snapshot=(master_sku.product_variant.product.category.name if master_sku else ""),
            subcategory_snapshot=(master_sku.product_variant.product.subcategory.name if master_sku and master_sku.product_variant.product.subcategory else ""),
            product_name_snapshot=(master_sku.product_variant.product.name if master_sku else ""),
            variant_name_snapshot=(master_sku.size if master_sku else ""),
            existing_line=existing_line,
            selected_source_data=selected_data,
        )
        staged_rows.append(staged)
        if is_out_of_scope and not historical_status_audit:
            out_of_scope_ids.add(staged.id)
        pending_issues.extend((staged, *issue) for issue in row_issues)

        if existing_line and not is_pure_cancelled and not is_out_of_scope:
            financial_changed = any(
                (
                    quantity != existing_line.quantity,
                    net_unit_price != existing_line.net_unit_price,
                )
            )
            if financial_changed:
                pending_issues.append(
                    (
                        staged,
                        "ERROR",
                        "IMMUTABLE_SNAPSHOT_CONFLICT",
                        "financial_snapshot",
                        "Qty/harga/COGS berbeda dari transaksi committed. Gunakan workflow koreksi, bukan overwrite import.",
                        True,
                    )
                )

    in_scope_rows = [row for row in staged_rows if row.id not in out_of_scope_ids]
    key_counts = Counter(row.business_key for row in in_scope_rows)
    duplicate_keys = {key for key, count in key_counts.items() if count > 1}
    for row in in_scope_rows:
        if row.business_key in duplicate_keys:
            pending_issues.append(
                (row, "ERROR", "DUPLICATE_BUSINESS_KEY", "business_key", f"Duplicate {row.business_key} dalam file.", True)
            )

    by_order = defaultdict(list)
    for row in in_scope_rows:
        by_order[row.order_number].append(row)
    for order_number, order_rows in by_order.items():
        statuses = {(row.normalized_status, row.is_pure_cancelled) for row in order_rows}
        datetimes = {row.order_datetime for row in order_rows}
        if len(statuses) > 1 or len(datetimes) > 1:
            for row in order_rows:
                pending_issues.append(
                    (
                        row,
                        "ERROR",
                        "ORDER_HEADER_CONFLICT",
                        "order",
                        f"Order {order_number} memiliki status/tanggal yang tidak konsisten antar-SKU.",
                        True,
                    )
                )

    issue_models = [_issue(batch, *issue) for issue in pending_issues]
    SalesImportIssue.objects.bulk_create(issue_models)
    blocked_ids = {issue[0].id for issue in pending_issues if issue[-1]}
    action_counts = Counter()
    updates = []
    for row in staged_rows:
        if row.id in out_of_scope_ids:
            action = StagedSalesRow.ProposedAction.OUT_OF_SCOPE
        elif row.id in blocked_ids:
            action = StagedSalesRow.ProposedAction.BLOCKED
        elif (
            row.selected_source_data.get("historical_status_audit")
            and not row.selected_source_data.get("historical_status_update_allowed", True)
        ):
            action = StagedSalesRow.ProposedAction.UNCHANGED
        elif row.is_pure_cancelled:
            action = StagedSalesRow.ProposedAction.PURE_CANCEL
        elif row.existing_line is None:
            action = StagedSalesRow.ProposedAction.NEW
        elif (
            row.existing_line.order.current_status != row.normalized_status
            or row.existing_line.order.source_status != row.source_status
        ):
            action = StagedSalesRow.ProposedAction.STATUS_UPDATE
        else:
            action = StagedSalesRow.ProposedAction.UNCHANGED
        row.proposed_action = action
        action_counts[action] += 1
        updates.append(row)
    StagedSalesRow.objects.bulk_update(updates, ["proposed_action"])

    blocking_count = sum(1 for issue in issue_models if issue.is_blocking)
    warning_count = sum(1 for issue in issue_models if issue.severity == "WARNING")
    batch.status = SalesImportBatch.Status.BLOCKED if blocking_count else SalesImportBatch.Status.READY
    batch.parser_version = PARSER_VERSION
    batch.total_rows = len(staged_rows)
    batch.new_rows = action_counts[StagedSalesRow.ProposedAction.NEW]
    batch.status_update_rows = action_counts[StagedSalesRow.ProposedAction.STATUS_UPDATE]
    batch.unchanged_rows = action_counts[StagedSalesRow.ProposedAction.UNCHANGED]
    batch.ignored_cancel_rows = action_counts[StagedSalesRow.ProposedAction.PURE_CANCEL]
    batch.out_of_scope_rows = action_counts[StagedSalesRow.ProposedAction.OUT_OF_SCOPE]
    batch.blocking_issue_count = blocking_count
    batch.warning_count = warning_count
    batch.data_start = min(dates) if dates else None
    batch.data_end = max(dates) if dates else None
    batch.previewed_at = timezone.now()
    batch.quality_summary = {
        **quality,
        "headers": headers,
        "duplicate_business_key_count": len(duplicate_keys),
        "sales_cutover_date": str(SALES_CUTOVER_DATE),
        "historical_status_audit_orders": len(historical_audit_order_numbers),
        "historical_evidence_only_rows": action_counts[StagedSalesRow.ProposedAction.OUT_OF_SCOPE],
        "historical_financial_snapshot_locked": True,
        "historical_inventory_posting": False,
    }
    batch.save()
    return batch
