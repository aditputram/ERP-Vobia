import csv
import math
import zipfile
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from master_data.models import MarketplaceProductMapping, SKU

from ..models import ImportValidationIssue, MasterImportBatch, StagedMasterRow


PARSER_VERSION = "master-v2"
MAX_SAFE_INTEGER = 9_007_199_254_740_991
CANONICAL_HEADERS = [
    "SOURCE",
    "SKU",
    "Parrent Sku",
    "ARTICLE",
    "CATEGORY",
    "SUB CATAGORY",
    "VARIANT",
    "SUB VARIANT",
    "STATUS PRODUCT",
    "COGS",
    "Retail Price",
    "Kode Shopee",
    "Kode Tiktok",
]


class MasterParseError(Exception):
    pass


def _text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        return format(value, ".15g")
    return str(value).strip()


def _json_value(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return _text(value)


def _decimal(value):
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        raise InvalidOperation
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip().replace("Rp", "").replace("rp", "").replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        tail = text.rsplit(",", maxsplit=1)[1]
        text = text.replace(",", "" if len(tail) == 3 else ".")
    elif "." in text:
        tail = text.rsplit(".", maxsplit=1)[1]
        if len(tail) == 3 and text.replace(".", "").replace("-", "").isdigit():
            text = text.replace(".", "")
    return Decimal(text)


def _identifier(value, field_name):
    issues = []
    if value in (None, ""):
        return "", issues
    if isinstance(value, bool):
        return "", [
            (
                "ERROR",
                "INVALID_IDENTIFIER_TYPE",
                field_name,
                "Kode marketplace tidak boleh berupa boolean.",
                True,
            )
        ]
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value) or not value.is_integer():
            return _text(value), [
                (
                    "ERROR",
                    "NUMERIC_IDENTIFIER_PRECISION_RISK",
                    field_name,
                    "Kode marketplace terbaca sebagai angka desimal dan presisinya tidak dapat dipercaya. Ekspor ulang sebagai text.",
                    True,
                )
            ]
        blocking = abs(value) > MAX_SAFE_INTEGER
        return str(int(value)), [
            (
                "ERROR" if blocking else "WARNING",
                "NUMERIC_IDENTIFIER_PRECISION_RISK" if blocking else "NUMERIC_IDENTIFIER_CONVERTED",
                field_name,
                (
                    "Kode marketplace melebihi batas integer aman dan harus berasal dari cell/text column."
                    if blocking
                    else "Kode marketplace sumber berupa angka; ERP mempertahankannya sebagai text."
                ),
                blocking,
            )
        ]
    if isinstance(value, int):
        blocking = field_name == "Kode Tiktok" and abs(value) > MAX_SAFE_INTEGER
        severity = "ERROR" if blocking else "WARNING"
        code = "NUMERIC_IDENTIFIER_PRECISION_RISK" if blocking else "NUMERIC_IDENTIFIER_CONVERTED"
        message = (
            "Kode TikTok melebihi batas integer aman dan harus berasal dari cell/text column."
            if blocking
            else "Kode marketplace sumber berupa angka; ERP mempertahankannya sebagai text."
        )
        issues.append((severity, code, field_name, message, blocking))
        return str(value), issues
    return str(value).strip(), issues


def _read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(handle, dialect=dialect)
        try:
            headers = [str(value).strip() for value in next(reader)]
        except StopIteration as exc:
            raise MasterParseError("File CSV kosong.") from exc
        rows = []
        for row_number, values in enumerate(reader, start=2):
            if not any(_text(value) for value in values):
                continue
            padded = list(values) + [None] * max(0, len(headers) - len(values))
            rows.append((row_number, dict(zip(headers, padded, strict=False))))
        return headers, rows, "CSV"


def _read_xlsx(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = next(
            (candidate for candidate in workbook.worksheets if candidate.title.strip().upper() == "VOBIA"),
            workbook.active,
        )
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = [_text(value) for value in next(iterator)]
        except StopIteration as exc:
            raise MasterParseError("File Excel kosong.") from exc
        rows = []
        for row_number, values in enumerate(iterator, start=2):
            if not any(_text(value) for value in values):
                continue
            padded = list(values) + [None] * max(0, len(headers) - len(values))
            rows.append((row_number, dict(zip(headers, padded, strict=False))))
        return headers, rows, sheet.title
    finally:
        workbook.close()


def _read_source(path, detected_format):
    if detected_format == "csv":
        return _read_csv(path)
    if detected_format == "xlsx":
        return _read_xlsx(path)
    raise MasterParseError("Format file tidak didukung.")


def _issue(batch, staged_row, severity, code, field_name, message, blocking):
    return ImportValidationIssue(
        batch=batch,
        staged_row=staged_row,
        severity=severity,
        code=code,
        field_name=field_name,
        message=message,
        is_blocking=blocking,
    )


def _existing_row_changed(staged, existing, existing_mappings):
    product = existing.product_variant.product
    subcategory_name = product.subcategory.name if product.subcategory else ""
    current = {
        "parent_sku": product.parent_sku,
        "article": product.article,
        "category": product.category.name,
        "subcategory": subcategory_name,
        "variant": existing.product_variant.name if existing.product_variant.name != "Default" else "",
        "sub_variant": existing.size,
        "product_status": product.status.name,
        "cogs": existing.current_master_cogs,
        "retail_price": existing.current_retail_price,
    }
    incoming = {
        "parent_sku": staged.parent_sku,
        "article": staged.article,
        "category": staged.category,
        "subcategory": staged.subcategory,
        "variant": staged.variant,
        "sub_variant": staged.sub_variant,
        "product_status": staged.product_status,
        "cogs": staged.cogs,
        "retail_price": staged.retail_price,
    }
    if current != incoming:
        return True

    mapping_key = product.id
    if staged.shopee_code and staged.shopee_code not in existing_mappings[(mapping_key, "Shopee")]:
        return True
    if staged.tiktok_code and staged.tiktok_code not in existing_mappings[(mapping_key, "Tiktok")]:
        return True
    return False


@transaction.atomic
def parse_master_batch(batch):
    batch = MasterImportBatch.objects.select_for_update().select_related("raw_file").get(pk=batch.pk)
    if batch.status != MasterImportBatch.Status.PARSING:
        raise MasterParseError("Batch hanya dapat diparsing satu kali.")

    path = Path(settings.PRIVATE_UPLOAD_ROOT) / batch.raw_file.storage_path
    try:
        headers, source_rows, source_sheet = _read_source(path, batch.raw_file.detected_format)
    except (MasterParseError, OSError, ValueError, zipfile.BadZipFile, InvalidFileException) as exc:
        ImportValidationIssue.objects.create(
            batch=batch,
            severity=ImportValidationIssue.Severity.ERROR,
            code="FILE_PARSE_ERROR",
            message=str(exc),
            is_blocking=True,
        )
        batch.status = MasterImportBatch.Status.BLOCKED
        batch.blocking_issue_count = 1
        batch.previewed_at = timezone.now()
        batch.save(update_fields=["status", "blocking_issue_count", "previewed_at"])
        return batch

    missing_headers = [header for header in CANONICAL_HEADERS if header not in headers]
    duplicate_headers = [name for name, count in Counter(headers).items() if name and count > 1]
    header_issues = []
    if missing_headers:
        header_issues.append(
            _issue(
                batch,
                None,
                "ERROR",
                "MISSING_CANONICAL_HEADERS",
                "header",
                "Header wajib tidak ditemukan: " + ", ".join(missing_headers),
                True,
            )
        )
    if duplicate_headers:
        header_issues.append(
            _issue(
                batch,
                None,
                "ERROR",
                "DUPLICATE_HEADERS",
                "header",
                "Header duplicate: " + ", ".join(duplicate_headers),
                True,
            )
        )
    if header_issues:
        ImportValidationIssue.objects.bulk_create(header_issues)
        batch.status = MasterImportBatch.Status.BLOCKED
        batch.blocking_issue_count = len(header_issues)
        batch.previewed_at = timezone.now()
        batch.quality_summary = {"source_sheet": source_sheet, "headers": headers}
        batch.save(
            update_fields=[
                "status",
                "blocking_issue_count",
                "previewed_at",
                "quality_summary",
            ]
        )
        return batch

    if len(source_rows) > settings.MASTER_IMPORT_MAX_ROWS:
        ImportValidationIssue.objects.create(
            batch=batch,
            severity=ImportValidationIssue.Severity.ERROR,
            code="ROW_LIMIT_EXCEEDED",
            message=f"File berisi {len(source_rows)} baris; batas aman saat ini {settings.MASTER_IMPORT_MAX_ROWS}.",
            is_blocking=True,
        )
        batch.status = MasterImportBatch.Status.BLOCKED
        batch.total_rows = len(source_rows)
        batch.blocking_issue_count = 1
        batch.previewed_at = timezone.now()
        batch.save(
            update_fields=["status", "total_rows", "blocking_issue_count", "previewed_at"]
        )
        return batch

    raw_skus = [_text(row.get("SKU")) for _, row in source_rows]
    existing_by_sku = {
        item.sku: item
        for item in SKU.objects.select_related(
            "product_variant__product__status",
            "product_variant__product__category",
            "product_variant__product__subcategory",
        ).filter(sku__in=raw_skus)
    }
    existing_product_ids = {
        item.product_variant.product_id for item in existing_by_sku.values()
    }
    existing_mappings = defaultdict(set)
    for mapping in MarketplaceProductMapping.objects.filter(product_id__in=existing_product_ids, is_active=True):
        existing_mappings[(mapping.product_id, mapping.source)].add(mapping.marketplace_product_code)

    staged_rows = []
    pending_issue_data = []
    quality = Counter()

    for row_number, raw in source_rows:
        cogs = None
        retail_price = None
        row_issue_data = []
        try:
            cogs = _decimal(raw.get("COGS"))
        except (InvalidOperation, ValueError):
            row_issue_data.append(("ERROR", "INVALID_DECIMAL", "COGS", "COGS bukan angka valid.", True))
        try:
            retail_price = _decimal(raw.get("Retail Price"))
        except (InvalidOperation, ValueError):
            row_issue_data.append(
                ("ERROR", "INVALID_DECIMAL", "Retail Price", "Retail Price bukan angka valid.", True)
            )

        shopee_code, shopee_issues = _identifier(raw.get("Kode Shopee"), "Kode Shopee")
        tiktok_code, tiktok_issues = _identifier(raw.get("Kode Tiktok"), "Kode Tiktok")
        row_issue_data.extend(shopee_issues)
        row_issue_data.extend(tiktok_issues)

        normalized = {
            "source": _text(raw.get("SOURCE")),
            "sku": _text(raw.get("SKU")),
            "parent_sku": _text(raw.get("Parrent Sku")),
            "article": _text(raw.get("ARTICLE")),
            "category": _text(raw.get("CATEGORY")),
            "subcategory": _text(raw.get("SUB CATAGORY")),
            "variant": _text(raw.get("VARIANT")),
            "sub_variant": _text(raw.get("SUB VARIANT")),
            "product_status": _text(raw.get("STATUS PRODUCT")),
        }

        required_fields = {
            "SOURCE": normalized["source"],
            "SKU": normalized["sku"],
            "ARTICLE": normalized["article"],
            "CATEGORY": normalized["category"],
            "STATUS PRODUCT": normalized["product_status"],
        }
        for field_name, value in required_fields.items():
            if not value:
                row_issue_data.append(
                    ("ERROR", "REQUIRED_VALUE_MISSING", field_name, f"{field_name} wajib diisi.", True)
                )

        if normalized["source"].lower() != "vobia":
            row_issue_data.append(
                ("ERROR", "INVALID_SOURCE", "SOURCE", "Master import ini hanya menerima SOURCE Vobia.", True)
            )
        if cogs is not None and cogs < 0:
            row_issue_data.append(("ERROR", "NEGATIVE_VALUE", "COGS", "COGS tidak boleh negatif.", True))
        if retail_price is not None and retail_price < 0:
            row_issue_data.append(
                ("ERROR", "NEGATIVE_VALUE", "Retail Price", "Retail Price tidak boleh negatif.", True)
            )
        if cogs is None:
            quality["missing_cogs"] += 1
            row_issue_data.append(("WARNING", "MISSING_COGS", "COGS", "COGS kosong; financial posting akan diblokir.", False))
        if retail_price is None:
            quality["missing_retail_price"] += 1
            row_issue_data.append(
                ("WARNING", "MISSING_RETAIL_PRICE", "Retail Price", "Retail Price kosong; sales posting akan diblokir.", False)
            )
        if not normalized["parent_sku"]:
            quality["missing_parent_sku"] += 1
            row_issue_data.append(
                ("WARNING", "MISSING_PARENT_SKU", "Parrent Sku", "Parent SKU kosong dan tidak akan ditebak.", False)
            )
        if not shopee_code:
            quality["missing_shopee_code"] += 1
            row_issue_data.append(
                (
                    "WARNING",
                    "MISSING_SHOPEE_CODE",
                    "Kode Shopee",
                    "Kode Shopee kosong; traffic Shopee tidak dapat dipetakan untuk record ini.",
                    False,
                )
            )
        if not tiktok_code:
            quality["missing_tiktok_code"] += 1
            row_issue_data.append(
                (
                    "WARNING",
                    "MISSING_TIKTOK_CODE",
                    "Kode Tiktok",
                    "Kode TikTok kosong; traffic TikTok tidak dapat dipetakan untuk record ini.",
                    False,
                )
            )

        existing = existing_by_sku.get(normalized["sku"])
        staged = StagedMasterRow.objects.create(
            batch=batch,
            row_number=row_number,
            cogs=cogs,
            retail_price=retail_price,
            shopee_code=shopee_code,
            tiktok_code=tiktok_code,
            existing_sku=existing,
            original_data={header: _json_value(raw.get(header)) for header in CANONICAL_HEADERS},
            **normalized,
        )
        staged_rows.append(staged)
        pending_issue_data.extend((staged, *item) for item in row_issue_data)

    counts = Counter(row.sku for row in staged_rows if row.sku)
    duplicate_skus = {sku for sku, count in counts.items() if count > 1}
    for staged in staged_rows:
        if staged.sku in duplicate_skus:
            pending_issue_data.append(
                (
                    staged,
                    "ERROR",
                    "DUPLICATE_SKU_IN_FILE",
                    "SKU",
                    f"SKU {staged.sku} muncul lebih dari satu kali dalam file.",
                    True,
                )
            )

    by_product = defaultdict(list)
    for staged in staged_rows:
        by_product[staged.product_key].append(staged)
    consistency_fields = ("category", "subcategory", "product_status")
    for product_key, rows in by_product.items():
        for field_name in consistency_fields:
            values = {getattr(row, field_name) for row in rows}
            if len(values) > 1:
                for staged in rows:
                    pending_issue_data.append(
                        (
                            staged,
                            "ERROR",
                            "PRODUCT_GROUP_CONFLICT",
                            field_name,
                            f"Product key {product_key} memiliki nilai {field_name} yang tidak konsisten.",
                            True,
                        )
                    )

    issues = [_issue(batch, *item) for item in pending_issue_data]
    ImportValidationIssue.objects.bulk_create(issues)
    blocked_row_ids = {item[0].id for item in pending_issue_data if item[-1]}

    action_counts = Counter()
    rows_to_update = []
    for staged in staged_rows:
        if staged.id in blocked_row_ids:
            action = StagedMasterRow.ProposedAction.BLOCKED
        elif staged.existing_sku is None:
            action = StagedMasterRow.ProposedAction.NEW
        elif _existing_row_changed(staged, staged.existing_sku, existing_mappings):
            action = StagedMasterRow.ProposedAction.UPDATE
        else:
            action = StagedMasterRow.ProposedAction.UNCHANGED
        staged.proposed_action = action
        action_counts[action] += 1
        rows_to_update.append(staged)
    StagedMasterRow.objects.bulk_update(rows_to_update, ["proposed_action"])

    blocking_count = sum(1 for issue in issues if issue.is_blocking)
    warning_count = sum(1 for issue in issues if issue.severity == "WARNING")
    batch.status = (
        MasterImportBatch.Status.BLOCKED if blocking_count else MasterImportBatch.Status.READY
    )
    batch.parser_version = PARSER_VERSION
    batch.total_rows = len(staged_rows)
    batch.new_rows = action_counts[StagedMasterRow.ProposedAction.NEW]
    batch.changed_rows = action_counts[StagedMasterRow.ProposedAction.UPDATE]
    batch.unchanged_rows = action_counts[StagedMasterRow.ProposedAction.UNCHANGED]
    batch.blocking_issue_count = blocking_count
    batch.warning_count = warning_count
    batch.previewed_at = timezone.now()
    batch.quality_summary = {
        **quality,
        "source_sheet": source_sheet,
        "headers": headers,
        "duplicate_sku_count": len(duplicate_skus),
    }
    batch.save()
    return batch
