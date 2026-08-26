import csv
import hashlib
import uuid
from calendar import monthrange
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from audit.services import record_audit
from imports.models import RawFile
from master_data.models import MarketplaceProductMapping, Product

from ..models import TrafficImportBatch, TrafficPeriodState, TrafficProductMetric


PARSER_VERSION = "traffic-history-v1"
WIDE_PARSER_VERSION = "traffic-history-wide-v1"
HISTORY_END = date(2026, 7, 31)


def _text(value):
    return "" if value is None else str(value).strip()


def _count(value):
    text = _text(value).replace(",", "").replace(".", "")
    try:
        result = int(text)
    except ValueError as exc:
        raise ValidationError(f"Metric traffic tidak valid: {value}") from exc
    if result < 0:
        raise ValidationError("Metric traffic tidak boleh negatif.")
    return result


def _month(value):
    number = int(_text(value).split(".", maxsplit=1)[0])
    return date(2026, number, 1)


def _headers(source):
    return {
        "code": "Kode Shopee" if source == "Shopee" else "Kode Tiktok",
        "product": "Produk" if source == "Shopee" else "Product",
        "category": "Category",
        "subcategory": "Sub Category",
        "views": "Jumlah Produk Dilihat",
        "clicks": "Produk Diklik",
        "visitors": "Pengunjung Produk (Kunjungan)",
    }


def _checksum(path):
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _wide_xlsx_rows(source_path, source):
    if source != "Tiktok":
        raise ValidationError("Historical traffic wide Excel hanya didukung untuk TikTok.")
    required = {
        "Product Name",
        "Jan Visitors",
        "Feb Visitors",
        "Jan Clicks",
        "Feb Clicks",
        "Jan Views",
        "Feb Views",
    }
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [_text(value) for value in next(rows, ())]
        missing = sorted(required - set(headers))
        if missing:
            raise ValidationError("Header historical traffic wide tidak lengkap: " + ", ".join(missing))
        positions = {header: headers.index(header) for header in required}
        products = defaultdict(list)
        for product in Product.objects.select_related("category", "subcategory"):
            products[product.name.strip().casefold()].append(product)
        codes_by_product = defaultdict(list)
        for mapping in MarketplaceProductMapping.objects.filter(source=source, is_active=True):
            codes_by_product[mapping.product_id].append(mapping.marketplace_product_code)

        rows_by_month = defaultdict(list)
        seen = {}
        for row_number, values in enumerate(rows, start=2):
            product_name = _text(values[positions["Product Name"]])
            if not product_name:
                continue
            metrics = tuple(
                0 if values[positions[header]] in (None, "") else _count(values[positions[header]])
                for header in (
                    "Jan Visitors",
                    "Feb Visitors",
                    "Jan Clicks",
                    "Feb Clicks",
                    "Jan Views",
                    "Feb Views",
                )
            )
            name_key = product_name.casefold()
            if name_key in seen:
                if seen[name_key] != metrics:
                    raise ValidationError(f"Duplicate Product Name dengan metric berbeda: {product_name}")
                continue
            seen[name_key] = metrics
            candidates = products[name_key]
            if len(candidates) != 1:
                raise ValidationError(f"Product Name harus cocok tepat ke satu Product canonical: {product_name}")
            product = candidates[0]
            codes = codes_by_product[product.id]
            if len(codes) > 1:
                raise ValidationError(f"Product memiliki lebih dari satu kode TikTok aktif: {product_name}")
            code = codes[0] if codes else ""
            traffic_key = f"PRODUCT::{product.id}"
            for month_number, visitors, clicks, views in (
                (1, metrics[0], metrics[2], metrics[4]),
                (2, metrics[1], metrics[3], metrics[5]),
            ):
                rows_by_month[date(2026, month_number, 1)].append(
                    {
                        "row_number": row_number,
                        "code": code,
                        "traffic_key": traffic_key,
                        "product": product,
                        "product_name": product_name,
                        "category": product.category.name,
                        "subcategory": product.subcategory.name if product.subcategory_id else "",
                        "views": views,
                        "clicks": clicks,
                        "visitors": visitors,
                    }
                )
        return rows_by_month
    finally:
        workbook.close()


@transaction.atomic
def migrate_historical_traffic(source_path, source, actor):
    if source not in {"Shopee", "Tiktok"}:
        raise ValidationError("Source historical traffic harus Shopee atau Tiktok.")
    source_path = Path(source_path)
    suffix = source_path.suffix.lower()
    if not source_path.is_file() or suffix not in {".csv", ".xlsx"}:
        raise ValidationError("Historical traffic source harus CSV atau Excel.")
    parser_version = WIDE_PARSER_VERSION if suffix == ".xlsx" else PARSER_VERSION
    checksum = _checksum(source_path)
    dataset = RawFile.DatasetType.TRAFFIC_SHOPEE if source == "Shopee" else RawFile.DatasetType.TRAFFIC_TIKTOK
    duplicate = RawFile.objects.filter(dataset_type=dataset, checksum_sha256=checksum).first()
    if duplicate:
        return list(duplicate.traffic_batches.filter(parser_version=parser_version).order_by("period_start"))

    rows_by_month = _wide_xlsx_rows(source_path, source) if suffix == ".xlsx" else defaultdict(list)
    key_counts = Counter()
    if suffix == ".csv":
        required = _headers(source)
        with source_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            missing = sorted(({"Month"} | set(required.values())) - set(reader.fieldnames or []))
            if missing:
                raise ValidationError("Header historical traffic tidak lengkap: " + ", ".join(missing))
            for row_number, raw in enumerate(reader, start=2):
                if not _text(raw.get("Month")):
                    continue
                month = _month(raw["Month"])
                if month > HISTORY_END:
                    continue
                code = _text(raw[required["code"]])
                traffic_key = code or f"NO_CODE::{row_number}"
                rows_by_month[month].append({
                    "row_number": row_number,
                    "code": code,
                    "traffic_key": traffic_key,
                    "product_name": _text(raw[required["product"]]),
                    "category": _text(raw[required["category"]]),
                    "subcategory": _text(raw[required["subcategory"]]),
                    "views": _count(raw[required["views"]]),
                    "clicks": _count(raw[required["clicks"]]),
                    "visitors": _count(raw[required["visitors"]]),
                })
    for month, rows in rows_by_month.items():
        for row in rows:
            key_counts[(month, row["traffic_key"])] += 1
    duplicates = [key for key, count in key_counts.items() if count > 1]
    if duplicates:
        raise ValidationError(f"Duplicate Source + Month + Product Code ditemukan: {len(duplicates)} key.")

    mappings = defaultdict(list)
    for mapping in MarketplaceProductMapping.objects.filter(source=source, is_active=True).select_related("product"):
        mappings[mapping.marketplace_product_code].append(mapping.product)

    current = timezone.localdate()
    relative = Path("traffic") / "historical" / source.lower() / str(current.year) / f"{current.month:02d}" / f"{uuid.uuid4()}{suffix}"
    destination = Path(settings.PRIVATE_UPLOAD_ROOT) / relative
    destination.parent.mkdir(parents=True, exist_ok=True)
    with source_path.open("rb") as source_file, destination.open("wb") as target:
        for chunk in iter(lambda: source_file.read(1024 * 1024), b""):
            target.write(chunk)

    raw = RawFile.objects.create(
        dataset_type=dataset,
        original_filename=source_path.name,
        storage_path=str(relative),
        checksum_sha256=checksum,
        byte_size=source_path.stat().st_size,
        detected_format=suffix.lstrip("."),
        uploaded_by=actor,
        source_metadata={
            "source": source,
            "dataset": "historical_product_traffic",
            "workbook": "Vobia Sales 2026",
            "sheet": f"Traffic Product {source}",
        },
    )
    batches = []
    now = timezone.now()
    try:
        for month, rows in sorted(rows_by_month.items()):
            period_end = date(month.year, month.month, monthrange(month.year, month.month)[1])
            ambiguous = 0
            unmapped = 0
            metrics = []
            batch = TrafficImportBatch.objects.create(
                raw_file=raw,
                source=source,
                period_start=month,
                period_end=period_end,
                parser_version=parser_version,
                status=TrafficImportBatch.Status.COMMITTED,
                total_rows=len(rows),
                ready_rows=len(rows),
                approved_by=actor,
                approved_at=now,
            )
            for row in rows:
                candidates = mappings.get(row["code"], [])
                product = row.get("product") or (candidates[0] if len(candidates) == 1 else None)
                ambiguous += int(not row.get("product") and len(candidates) > 1)
                unmapped += int(product is None)
                metrics.append(TrafficProductMetric(
                    source=source,
                    period_start=month,
                    period_end=period_end,
                    product=product,
                    traffic_product_key=row["traffic_key"],
                    marketplace_product_code_snapshot=row["code"],
                    product_name_snapshot=row["product_name"],
                    category_snapshot=row["category"],
                    subcategory_snapshot=row["subcategory"],
                    is_historical_migration=True,
                    views=row["views"],
                    clicks=row["clicks"],
                    visitors=row["visitors"],
                    source_batch=batch,
                ))
            TrafficProductMetric.objects.bulk_create(metrics, batch_size=1000)
            batch.quality_summary = {
                "historical_migration": True,
                "ambiguous_current_product_mapping": ambiguous,
                "unmapped_current_product_code": unmapped,
                "inventory_posting": False,
                "views": sum(item.views for item in metrics),
                "clicks": sum(item.clicks for item in metrics),
                "visitors": sum(item.visitors for item in metrics),
            }
            batch.warning_count = ambiguous + unmapped
            batch.save(update_fields=["quality_summary", "warning_count"])
            TrafficPeriodState.objects.update_or_create(
                source=source,
                month=month,
                defaults={
                    "is_complete": True,
                    "last_successful_import_at": now,
                    "last_data_end": period_end,
                },
            )
            batches.append(batch)
        record_audit(
            actor=actor,
            action="historical_traffic_migration_committed",
            entity_type="imports.rawfile",
            entity_id=raw.id,
            after_values={
                "source": source,
                "months": len(batches),
                "metrics": sum(batch.ready_rows for batch in batches),
                "views": sum(batch.quality_summary["views"] for batch in batches),
                "clicks": sum(batch.quality_summary["clicks"] for batch in batches),
                "visitors": sum(batch.quality_summary["visitors"] for batch in batches),
            },
            metadata={"checksum_sha256": checksum, "parser_version": parser_version},
        )
    except Exception:
        destination.unlink(missing_ok=True)
        raise
    return batches
